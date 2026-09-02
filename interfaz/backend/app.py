from flask import Flask, render_template, request, jsonify, send_file
from sqlalchemy import text
import pandas as pd
from db import engine
from io import BytesIO
import subprocess
import sys
import threading
import uuid
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from dotenv import load_dotenv

PROJECT_ROOT = Path(__file__).resolve().parents[2]
load_dotenv(PROJECT_ROOT / ".env")
load_dotenv(Path(__file__).resolve().parent / ".env")


app = Flask(__name__)

planificacion_executor = ThreadPoolExecutor(max_workers=1)
planificacion_trabajos = {}
planificacion_trabajos_lock = threading.Lock()


def ejecutar_generacion_planificacion(job_id, comando, ruta_script, fecha_inicio, fecha_fin):
    try:
        resultado = subprocess.run(
            comando,
            capture_output=True,
            text=True,
            cwd=str(ruta_script.parent)
        )
        filas_guardadas = None
        version_guardada = None
        if resultado.returncode == 0:
            with engine.begin() as conn:
                version_guardada = conn.execute(text("""
                    SELECT id_version
                    FROM planificacion_version
                    WHERE activa = 1
                    ORDER BY id_version DESC
                    LIMIT 1
                """)).scalar()
                if version_guardada is not None:
                    filas_guardadas = conn.execute(text("""
                        SELECT COUNT(*)
                        FROM calendarizacion
                        WHERE id_version = :id_version
                          AND fecha BETWEEN :fecha_inicio AND :fecha_fin
                    """), {
                        "id_version": version_guardada,
                        "fecha_inicio": fecha_inicio,
                        "fecha_fin": fecha_fin,
                    }).scalar()
        with planificacion_trabajos_lock:
            planificacion_trabajos[job_id] = {
                "estado": "completado" if resultado.returncode == 0 else "error",
                "codigo_salida": resultado.returncode,
                "filas_guardadas": filas_guardadas,
                "version_guardada": version_guardada,
                "salida": resultado.stdout[-8000:],
                "detalle": resultado.stderr[-4000:] if resultado.returncode != 0 else "",
            }
    except Exception as error:
        with planificacion_trabajos_lock:
            planificacion_trabajos[job_id] = {
                "estado": "error",
                "detalle": str(error),
            }


def ensure_prophet_config_table():
    with engine.begin() as conn:
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS configuracion_prophet (
                id_centro INT NOT NULL DEFAULT 1,
                yearly_seasonality INT NOT NULL DEFAULT 20,
                weekly_seasonality INT NOT NULL DEFAULT 3,
                daily_seasonality TINYINT NOT NULL DEFAULT 0,
                seasonality_mode VARCHAR(30) NOT NULL DEFAULT 'multiplicative',
                interval_width DECIMAL(4,2) NOT NULL DEFAULT 0.80,
                n_changepoints INT NOT NULL DEFAULT 50,
                tasa_crecimiento DECIMAL(6,4) NOT NULL DEFAULT 0.0000,
                fecha_actualizacion DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                PRIMARY KEY (id_centro)
            )
        """))

        conn.execute(text("""
            INSERT INTO configuracion_prophet (
                id_centro,
                yearly_seasonality,
                weekly_seasonality,
                daily_seasonality,
                seasonality_mode,
                interval_width,
                n_changepoints,
                tasa_crecimiento
            )
            SELECT 1, 20, 3, 0, 'multiplicative', 0.80, 50, 0.0000
            WHERE NOT EXISTS (
                SELECT 1
                FROM configuracion_prophet
                WHERE id_centro = 1
            )
        """))


def get_prophet_config():
    ensure_prophet_config_table()

    with engine.begin() as conn:
        row = conn.execute(text("""
            SELECT
                yearly_seasonality,
                weekly_seasonality,
                daily_seasonality,
                seasonality_mode,
                interval_width,
                n_changepoints,
                tasa_crecimiento
            FROM configuracion_prophet
            WHERE id_centro = 1
        """)).mappings().first()

    if row is None:
        return {
            "yearly_seasonality": 20,
            "weekly_seasonality": 3,
            "daily_seasonality": False,
            "seasonality_mode": "multiplicative",
            "interval_width": 0.8,
            "n_changepoints": 50,
            "tasa_crecimiento": 0.0,
        }

    return {
        "yearly_seasonality": int(row["yearly_seasonality"]),
        "weekly_seasonality": int(row["weekly_seasonality"]),
        "daily_seasonality": bool(int(row["daily_seasonality"])),
        "seasonality_mode": row["seasonality_mode"] or "multiplicative",
        "interval_width": float(row["interval_width"]),
        "n_changepoints": int(row["n_changepoints"]),
        "tasa_crecimiento": float(row["tasa_crecimiento"]),
    }


def save_prophet_config(data):
    payload = data or {}

    defaults = get_prophet_config()

    yearly_seasonality = int(payload.get("yearly_seasonality", defaults["yearly_seasonality"]))
    weekly_seasonality = int(payload.get("weekly_seasonality", defaults["weekly_seasonality"]))
    daily_raw = payload.get("daily_seasonality", defaults["daily_seasonality"])
    daily_seasonality = 1 if str(daily_raw).lower() in {"1", "true", "yes", "verdadero"} else 0
    seasonality_mode = str(payload.get("seasonality_mode", defaults["seasonality_mode"]))
    if seasonality_mode not in {"additive", "multiplicative"}:
        seasonality_mode = "multiplicative"
    interval_width = float(payload.get("interval_width", defaults["interval_width"]))
    n_changepoints = int(payload.get("n_changepoints", defaults["n_changepoints"]))
    tasa_crecimiento = float(payload.get("tasa_crecimiento", defaults["tasa_crecimiento"]))

    with engine.begin() as conn:
        conn.execute(text("""
            INSERT INTO configuracion_prophet (
                id_centro,
                yearly_seasonality,
                weekly_seasonality,
                daily_seasonality,
                seasonality_mode,
                interval_width,
                n_changepoints,
                tasa_crecimiento
            ) VALUES (
                1,
                :yearly_seasonality,
                :weekly_seasonality,
                :daily_seasonality,
                :seasonality_mode,
                :interval_width,
                :n_changepoints,
                :tasa_crecimiento
            )
            ON DUPLICATE KEY UPDATE
                yearly_seasonality = VALUES(yearly_seasonality),
                weekly_seasonality = VALUES(weekly_seasonality),
                daily_seasonality = VALUES(daily_seasonality),
                seasonality_mode = VALUES(seasonality_mode),
                interval_width = VALUES(interval_width),
                n_changepoints = VALUES(n_changepoints),
                tasa_crecimiento = VALUES(tasa_crecimiento),
                fecha_actualizacion = CURRENT_TIMESTAMP
        """), {
            "yearly_seasonality": yearly_seasonality,
            "weekly_seasonality": weekly_seasonality,
            "daily_seasonality": daily_seasonality,
            "seasonality_mode": seasonality_mode,
            "interval_width": interval_width,
            "n_changepoints": n_changepoints,
            "tasa_crecimiento": tasa_crecimiento,
        })

    return get_prophet_config()


def ensure_promocion_tier_column():
    with engine.begin() as conn:
        cols = conn.execute(text("""
            SHOW COLUMNS FROM promocion LIKE 'tier'
        """)).fetchall()

        if not cols:
            conn.execute(text("""
                ALTER TABLE promocion
                ADD COLUMN tier INT NOT NULL DEFAULT 3 AFTER id_centro
            """))


def ensure_optimizacion_config_table():
    with engine.begin() as conn:
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS configuracion_optimizacion (
                id_centro INT NOT NULL DEFAULT 1,
                horas_presencia_mostrador DECIMAL(10,4) NOT NULL DEFAULT 11.0000,
                horas_otras_gestiones DECIMAL(10,4) NOT NULL DEFAULT 1.0000,
                porcentaje_devoluciones DECIMAL(10,4) NOT NULL DEFAULT 0.0500,
                horas_gestion_devoluciones DECIMAL(10,4) NOT NULL DEFAULT 3.0000,
                recoleccion_1_lineas INT NOT NULL DEFAULT 1000,
                recoleccion_1_tiempo_min DECIMAL(10,4) NOT NULL DEFAULT 1.8000,
                recoleccion_2_lineas INT NOT NULL DEFAULT 1000,
                recoleccion_2_tiempo_min DECIMAL(10,4) NOT NULL DEFAULT 0.3150,
                empaquetado_lineas INT NOT NULL DEFAULT 1000,
                empaquetado_tiempo_min DECIMAL(10,4) NOT NULL DEFAULT 5.0400,
                almacenado_lineas INT NOT NULL DEFAULT 1000,
                almacenado_tiempo_min DECIMAL(10,4) NOT NULL DEFAULT 4.3500,
                entrega_lineas INT NOT NULL DEFAULT 1000,
                entrega_tiempo_min DECIMAL(10,4) NOT NULL DEFAULT 2.9800,
                fecha_actualizacion DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                PRIMARY KEY (id_centro)
            )
        """))

        conn.execute(text("""
            INSERT INTO configuracion_optimizacion (
                id_centro,
                horas_presencia_mostrador,
                horas_otras_gestiones,
                porcentaje_devoluciones,
                horas_gestion_devoluciones,
                recoleccion_1_lineas,
                recoleccion_1_tiempo_min,
                recoleccion_2_lineas,
                recoleccion_2_tiempo_min,
                empaquetado_lineas,
                empaquetado_tiempo_min,
                almacenado_lineas,
                almacenado_tiempo_min,
                entrega_lineas,
                entrega_tiempo_min
            )
            SELECT 1,
                   11.0000,
                   1.0000,
                   0.0500,
                   3.0000,
                   1000,
                   1.8000,
                   1000,
                   0.3150,
                   1000,
                   5.0400,
                   1000,
                   4.3500,
                   1000,
                   2.9800
            WHERE NOT EXISTS (
                SELECT 1 FROM configuracion_optimizacion WHERE id_centro = 1
            )
        """))


def get_optimizacion_config():
    ensure_optimizacion_config_table()

    with engine.begin() as conn:
        row = conn.execute(text("""
            SELECT
                horas_presencia_mostrador,
                horas_otras_gestiones,
                porcentaje_devoluciones,
                horas_gestion_devoluciones,
                recoleccion_1_lineas,
                recoleccion_1_tiempo_min,
                recoleccion_2_lineas,
                recoleccion_2_tiempo_min,
                empaquetado_lineas,
                empaquetado_tiempo_min,
                almacenado_lineas,
                almacenado_tiempo_min,
                entrega_lineas,
                entrega_tiempo_min
            FROM configuracion_optimizacion
            WHERE id_centro = 1
        """)).mappings().first()

    if row is None:
        return {
            "horas_presencia_mostrador": 11.0,
            "horas_otras_gestiones": 1.0,
            "porcentaje_devoluciones": 0.05,
            "horas_gestion_devoluciones": 3.0,
            "recoleccion_1_lineas": 1000,
            "recoleccion_1_tiempo_min": 1.8,
            "recoleccion_2_lineas": 1000,
            "recoleccion_2_tiempo_min": 0.315,
            "empaquetado_lineas": 1000,
            "empaquetado_tiempo_min": 5.04,
            "almacenado_lineas": 1000,
            "almacenado_tiempo_min": 4.35,
            "entrega_lineas": 1000,
            "entrega_tiempo_min": 2.98,
        }

    return {
        "horas_presencia_mostrador": float(row["horas_presencia_mostrador"]),
        "horas_otras_gestiones": float(row["horas_otras_gestiones"]),
        "porcentaje_devoluciones": float(row["porcentaje_devoluciones"]),
        "horas_gestion_devoluciones": float(row["horas_gestion_devoluciones"]),
        "recoleccion_1_lineas": int(row["recoleccion_1_lineas"]),
        "recoleccion_1_tiempo_min": float(row["recoleccion_1_tiempo_min"]),
        "recoleccion_2_lineas": int(row["recoleccion_2_lineas"]),
        "recoleccion_2_tiempo_min": float(row["recoleccion_2_tiempo_min"]),
        "empaquetado_lineas": int(row["empaquetado_lineas"]),
        "empaquetado_tiempo_min": float(row["empaquetado_tiempo_min"]),
        "almacenado_lineas": int(row["almacenado_lineas"]),
        "almacenado_tiempo_min": float(row["almacenado_tiempo_min"]),
        "entrega_lineas": int(row["entrega_lineas"]),
        "entrega_tiempo_min": float(row["entrega_tiempo_min"]),
    }


def save_optimizacion_config(data):
    payload = data or {}
    defaults = get_optimizacion_config()

    params = {
        "horas_presencia_mostrador": float(payload.get("horas_presencia_mostrador", defaults["horas_presencia_mostrador"])),
        "horas_otras_gestiones": float(payload.get("horas_otras_gestiones", defaults["horas_otras_gestiones"])),
        "porcentaje_devoluciones": float(payload.get("porcentaje_devoluciones", defaults["porcentaje_devoluciones"])),
        "horas_gestion_devoluciones": float(payload.get("horas_gestion_devoluciones", defaults["horas_gestion_devoluciones"])),
        "recoleccion_1_lineas": int(payload.get("recoleccion_1_lineas", defaults["recoleccion_1_lineas"])),
        "recoleccion_1_tiempo_min": float(payload.get("recoleccion_1_tiempo_min", defaults["recoleccion_1_tiempo_min"])),
        "recoleccion_2_lineas": int(payload.get("recoleccion_2_lineas", defaults["recoleccion_2_lineas"])),
        "recoleccion_2_tiempo_min": float(payload.get("recoleccion_2_tiempo_min", defaults["recoleccion_2_tiempo_min"])),
        "empaquetado_lineas": int(payload.get("empaquetado_lineas", defaults["empaquetado_lineas"])),
        "empaquetado_tiempo_min": float(payload.get("empaquetado_tiempo_min", defaults["empaquetado_tiempo_min"])),
        "almacenado_lineas": int(payload.get("almacenado_lineas", defaults["almacenado_lineas"])),
        "almacenado_tiempo_min": float(payload.get("almacenado_tiempo_min", defaults["almacenado_tiempo_min"])),
        "entrega_lineas": int(payload.get("entrega_lineas", defaults["entrega_lineas"])),
        "entrega_tiempo_min": float(payload.get("entrega_tiempo_min", defaults["entrega_tiempo_min"])),
    }

    with engine.begin() as conn:
        conn.execute(text("""
            INSERT INTO configuracion_optimizacion (
                id_centro,
                horas_presencia_mostrador,
                horas_otras_gestiones,
                porcentaje_devoluciones,
                horas_gestion_devoluciones,
                recoleccion_1_lineas,
                recoleccion_1_tiempo_min,
                recoleccion_2_lineas,
                recoleccion_2_tiempo_min,
                empaquetado_lineas,
                empaquetado_tiempo_min,
                almacenado_lineas,
                almacenado_tiempo_min,
                entrega_lineas,
                entrega_tiempo_min
            ) VALUES (
                1,
                :horas_presencia_mostrador,
                :horas_otras_gestiones,
                :porcentaje_devoluciones,
                :horas_gestion_devoluciones,
                :recoleccion_1_lineas,
                :recoleccion_1_tiempo_min,
                :recoleccion_2_lineas,
                :recoleccion_2_tiempo_min,
                :empaquetado_lineas,
                :empaquetado_tiempo_min,
                :almacenado_lineas,
                :almacenado_tiempo_min,
                :entrega_lineas,
                :entrega_tiempo_min
            )
            ON DUPLICATE KEY UPDATE
                horas_presencia_mostrador = VALUES(horas_presencia_mostrador),
                horas_otras_gestiones = VALUES(horas_otras_gestiones),
                porcentaje_devoluciones = VALUES(porcentaje_devoluciones),
                horas_gestion_devoluciones = VALUES(horas_gestion_devoluciones),
                recoleccion_1_lineas = VALUES(recoleccion_1_lineas),
                recoleccion_1_tiempo_min = VALUES(recoleccion_1_tiempo_min),
                recoleccion_2_lineas = VALUES(recoleccion_2_lineas),
                recoleccion_2_tiempo_min = VALUES(recoleccion_2_tiempo_min),
                empaquetado_lineas = VALUES(empaquetado_lineas),
                empaquetado_tiempo_min = VALUES(empaquetado_tiempo_min),
                almacenado_lineas = VALUES(almacenado_lineas),
                almacenado_tiempo_min = VALUES(almacenado_tiempo_min),
                entrega_lineas = VALUES(entrega_lineas),
                entrega_tiempo_min = VALUES(entrega_tiempo_min),
                fecha_actualizacion = CURRENT_TIMESTAMP
        """), params)

    return get_optimizacion_config()


ensure_promocion_tier_column()


def get_all_config():
    return {
        **get_prophet_config(),
        **get_optimizacion_config(),
    }


def save_all_config(data):
    payload = data or {}

    prophet_keys = {
        "yearly_seasonality",
        "weekly_seasonality",
        "daily_seasonality",
        "seasonality_mode",
        "interval_width",
        "n_changepoints",
        "tasa_crecimiento",
    }

    prophet_payload = {key: payload[key] for key in sorted(payload.keys()) if key in prophet_keys}
    optimizacion_payload = {key: payload[key] for key in sorted(payload.keys()) if key not in prophet_keys}

    prophet_result = save_prophet_config(prophet_payload) if prophet_payload else get_prophet_config()
    optimizacion_result = save_optimizacion_config(optimizacion_payload) if optimizacion_payload else get_optimizacion_config()

    return {
        **prophet_result,
        **optimizacion_result,
    }


# ============================================================
# DASHBOARD
# ============================================================

@app.route("/dashboard")
def dashboard():
    trabajadores = pd.read_sql("""
        SELECT COUNT(*) AS total
        FROM trabajador
        WHERE activo = 1
    """, engine).iloc[0]["total"]

    pedidos = pd.read_sql("""
        SELECT SUM(total_pedidos) AS total
        FROM pedidohistorico
        WHERE id_centro = 1
    """, engine).iloc[0]["total"]

    ultima_prediccion = pd.read_sql("""
        SELECT MAX(fecha_generacion) AS fecha
        FROM prediccion
        WHERE id_centro = 1
    """, engine).iloc[0]["fecha"]

    ultima_planificacion = pd.read_sql("""
        SELECT MAX(fecha_generacion) AS fecha
        FROM planificacion
        WHERE id_centro = 1
    """, engine).iloc[0]["fecha"]

    return render_template(
        "dashboard.html",
        trabajadores=int(trabajadores or 0),
        pedidos=int(pedidos or 0),
        ultima_prediccion=ultima_prediccion,
        ultima_planificacion=ultima_planificacion
    )


# ============================================================
# GESTIÓN DE DATOS — TRABAJADORES (listado y consultas)
# ============================================================

# ---- Vista principal: trabajadores, contratos, tareas y disponibilidad ----
@app.route("/gestion-datos")
def gestion_datos():

    trabajadores = pd.read_sql("""

        SELECT

            t.id_trabajador,
            t.numero_vendedor,
            t.nombre,
            t.apellidos,
            t.correo,
            t.id_contrato,
            t.id_centro,
            t.fijo_discontinuo,
            t.disponibilidad,
            t.estado,
            t.activo,

            c.nombre AS contrato,

            COUNT(tt.id_tarea) AS competencias,

            GROUP_CONCAT(
                ta.nombre
                ORDER BY ta.nombre
                SEPARATOR ', '
            ) AS lista_competencias

        FROM trabajador t

        LEFT JOIN contrato c
            ON t.id_contrato = c.id_contrato

        LEFT JOIN trabajador_tarea tt
            ON t.id_trabajador = tt.id_trabajador

        LEFT JOIN tarea ta
            ON tt.id_tarea = ta.id_tarea

        WHERE t.activo = 1

        GROUP BY

            t.id_trabajador,
            t.numero_vendedor,
            t.nombre,
            t.apellidos,
            t.correo,
            t.id_contrato,
            t.id_centro,
            t.fijo_discontinuo,
            t.disponibilidad,
            t.estado,
            t.activo,
            c.nombre

        ORDER BY t.apellidos, t.nombre

        """, engine)

    contratos = pd.read_sql("""

        SELECT

            c.id_contrato,
            c.nombre,
            c.jornada,
            c.horas_anuales,
            c.horas_por_turno,

            COUNT(t.id_trabajador) AS trabajadores

        FROM contrato c

        LEFT JOIN trabajador t

            ON c.id_contrato = t.id_contrato
            AND t.activo = 1

        GROUP BY

            c.id_contrato,
            c.nombre,
            c.jornada,
            c.horas_anuales,
            c.horas_por_turno

        ORDER BY c.jornada DESC

        """, engine)


    tareas = pd.read_sql("""

        SELECT

            t.id_tarea,
            t.nombre,
            t.descripcion,
            t.activa,

            COUNT(tt.id_trabajador) AS trabajadores

        FROM tarea t

        LEFT JOIN trabajador_tarea tt
            ON t.id_tarea = tt.id_tarea

        GROUP BY
            t.id_tarea,
            t.nombre,
            t.descripcion,
            t.activa

        ORDER BY t.nombre

    """, engine)

    turnos = pd.read_sql("""
        SELECT *
        FROM turno_vacaciones
        ORDER BY codigo
    """, engine)

    total_tareas = pd.read_sql("""

        SELECT COUNT(*) AS total

        FROM tarea

        """, engine).iloc[0]["total"]


    disponibilidades=pd.read_sql("""
        SELECT
            t.id_trabajador,
            CONCAT(t.nombre,' ',t.apellidos) AS trabajador,
            COUNT(d.id_disponibilidad) AS total_periodos
        FROM trabajador t
        LEFT JOIN disponibilidad d
            ON t.id_trabajador=d.id_trabajador
        WHERE t.activo=1
        GROUP BY
            t.id_trabajador,
            t.nombre,
            t.apellidos
        ORDER BY
            t.apellidos,
            t.nombre
    """,engine)
        
    return render_template(
        "gestion-datos.html",
        trabajadores=trabajadores.to_dict(orient="records"),
        contratos=contratos.to_dict(orient="records"),
        tareas=tareas.to_dict(orient="records"),
        turnos=turnos.to_dict(orient="records"),
        total_tareas=total_tareas,
        disponibilidades=disponibilidades.to_dict(orient="records")
    )


# ---- Competencias de un trabajador (AJAX) ----
@app.route("/competencias_trabajador/<int:id_trabajador>")
def competencias_trabajador(id_trabajador):

    competencias = pd.read_sql("""

        SELECT

            ta.id_tarea,
            ta.nombre

        FROM trabajador_tarea tt

        JOIN tarea ta

            ON ta.id_tarea = tt.id_tarea

        WHERE tt.id_trabajador = %s

        ORDER BY ta.nombre

    """, engine, params=(id_trabajador,))

    return competencias.to_json(
        orient="records",
        force_ascii=False
    )


# ---- Trabajadores asociados a un contrato (AJAX) ----
@app.route("/trabajadores_contrato/<int:id_contrato>")
def trabajadores_contrato(id_contrato):

    trabajadores = pd.read_sql(

        text("""

            SELECT

                CONCAT(nombre,' ',apellidos) AS nombre

            FROM trabajador

            WHERE id_contrato = :id
            AND activo = 1

            ORDER BY apellidos, nombre

        """),

        engine,

        params={
            "id": id_contrato
        }

    )

    return jsonify(
        trabajadores.to_dict(orient="records")
    )


# ============================================================
# GESTIÓN DE DATOS — CONTRATOS
# ============================================================

# ---- Nuevo contrato ----
@app.route("/nuevo_contrato", methods=["POST"])
def nuevo_contrato():

    datos = request.get_json()

    with engine.begin() as conn:

        conn.execute(text("""

            INSERT INTO contrato(

                nombre,
                jornada,
                horas_anuales,
                horas_por_turno

            )

            VALUES(

                :nombre,
                :jornada,
                :horas_anuales,
                :horas_por_turno

            )

        """),

        {

            "nombre": datos["nombre"],
            "jornada": datos["jornada"],
            "horas_anuales": datos["horas_anuales"],
            "horas_por_turno": datos["horas_por_turno"]

        })

    return "",204


# ---- Eliminar contrato ----
@app.route("/eliminar_contrato", methods=["POST"])
def eliminar_contrato():

    datos = request.get_json()

    with engine.begin() as conn:

        trabajadores = conn.execute(text("""

            SELECT COUNT(*)

            FROM trabajador

            WHERE id_contrato = :id
            AND activo = 1

        """),

        {

            "id": datos["id"]

        }).scalar()

        if trabajadores > 0:

            return jsonify({

                "ok": False,

                "error":
                "No se puede eliminar este contrato porque tiene trabajadores asignados."

            })

        conn.execute(text("""

            DELETE

            FROM contrato

            WHERE id_contrato = :id

        """),

        {

            "id": datos["id"]

        })

    return jsonify({

        "ok": True

    })


# ---- Editar contrato ----
@app.route("/editar_contrato", methods=["POST"])
def editar_contrato():

    datos = request.get_json()

    with engine.begin() as conn:

        conn.execute(text("""

            UPDATE contrato

            SET

                nombre = :nombre,
                jornada = :jornada,
                horas_anuales = :horas_anuales,
                horas_por_turno = :horas_por_turno

            WHERE id_contrato = :id

        """),

        {

            "id": datos["id"],
            "nombre": datos["nombre"],
            "jornada": datos["jornada"],
            "horas_anuales": datos["horas_anuales"],
            "horas_por_turno": datos["horas_por_turno"]

        })

    return "",204


# ============================================================
# GESTIÓN DE DATOS — TRABAJADORES (alta, edición y estado)
# ============================================================

# ---- Eliminar trabajador ----
@app.route("/eliminar_trabajador", methods=["POST"])
def eliminar_trabajador():

    datos = request.get_json()
    print(datos)

    with engine.begin() as conn:

        conn.execute(text("""
            UPDATE trabajador
            SET activo=0
            WHERE id_trabajador=:id
        """),{"id":datos["id"]})

    return jsonify(ok=True)


# ---- Nuevo trabajador ----
@app.route("/nuevo_trabajador",methods=["POST"])
def nuevo_trabajador():
    datos=request.get_json()
    with engine.begin() as conn:
        resultado=conn.execute(text("""
            INSERT INTO trabajador(
                numero_vendedor,
                nombre,
                apellidos,
                correo,
                activo,
                id_contrato,
                id_centro,
                estado,
                disponibilidad,
                fijo_discontinuo
            )
            VALUES(
                :numero_vendedor,
                :nombre,
                :apellidos,
                :correo,
                1,
                :id_contrato,
                1,
                :estado,
                :disponibilidad,
                :fijo_discontinuo
            )
        """),datos)
        id_trabajador=resultado.lastrowid
        for tarea in datos["tareas"]:
            conn.execute(text("""
                INSERT INTO trabajador_tarea(
                    id_trabajador,
                    id_tarea
                )
                VALUES(
                    :id_trabajador,
                    :id_tarea
                )
            """),{
                "id_trabajador":id_trabajador,
                "id_tarea":tarea
            })
    return jsonify(ok=True)


# ---- Editar trabajador ----
@app.route("/editar_trabajador", methods=["POST"])
def editar_trabajador():

    datos = request.json

    with engine.begin() as conn:

        conn.execute(text("""

            UPDATE trabajador
            SET
                numero_vendedor=:numero_vendedor,
                nombre=:nombre,
                apellidos=:apellidos,
                correo=:correo,
                id_contrato=:id_contrato,
                disponibilidad=:disponibilidad,
                estado=:estado,
                fijo_discontinuo=:fijo_discontinuo
            WHERE id_trabajador=:id

        """), datos)

        # Eliminar competencias anteriores
        conn.execute(text("""

            DELETE FROM trabajador_tarea

            WHERE id_trabajador = :id

        """), {"id": datos["id"]})

        # Insertar las nuevas
        for id_tarea in datos["tareas"]:

            conn.execute(text("""

                INSERT INTO trabajador_tarea
                (id_trabajador, id_tarea)

                VALUES
                (:id_trabajador, :id_tarea)

            """), {

                "id_trabajador": datos["id"],
                "id_tarea": id_tarea

            })

    return jsonify(ok=True)


# ---- Cambiar estado (activo / inactivo) ----
@app.route("/cambiar_estado_trabajador", methods=["POST"])
def cambiar_estado_trabajador():

    datos = request.json

    with engine.begin() as conn:

        conn.execute(text("""

            UPDATE trabajador

            SET estado = NOT estado

            WHERE id_trabajador = :id

        """),{

            "id":datos["id"]

        })

    return jsonify(ok=True)


# ---- Periodos de fijo discontinuo de un trabajador (AJAX) ----
@app.route("/periodos_trabajador/<int:id_trabajador>")
def periodos_trabajador(id_trabajador):
    periodos=pd.read_sql("""
        SELECT
            id_periodo,
            fecha_inicio,
            fecha_fin
        FROM periodo_fijo_discontinuo
        WHERE id_trabajador=%s
        ORDER BY fecha_inicio
    """,engine,params=(id_trabajador,))
    return periodos.to_json(
        orient="records",
        date_format="iso"
    )


# ---- Nuevo periodo de fijo discontinuo ----
@app.route("/nuevo_periodo_fd",methods=["POST"])
def nuevo_periodo_fd():
    datos=request.get_json()
    with engine.begin() as conn:
        conn.execute(text("""
            INSERT INTO periodo_fijo_discontinuo(
                id_trabajador,
                fecha_inicio,
                fecha_fin
            )
            VALUES(
                :id_trabajador,
                :fecha_inicio,
                :fecha_fin
            )
        """),datos)
    return jsonify(ok=True)


# ---- Editar periodo de fijo discontinuo ----
@app.route("/editar_periodo_fd",methods=["POST"])
def editar_periodo_fd():
    datos=request.get_json()
    with engine.begin() as conn:
        conn.execute(text("""
            UPDATE periodo_fijo_discontinuo
            SET
                fecha_inicio=:fecha_inicio,
                fecha_fin=:fecha_fin
            WHERE id_periodo=:id_periodo
        """),datos)
    return jsonify(ok=True)


# ---- Eliminar periodo de fijo discontinuo ----
@app.route("/eliminar_periodo_fd",methods=["POST"])
def eliminar_periodo_fd():
    datos=request.get_json()
    print("ELIMINANDO PERIODO:",datos)
    with engine.begin() as conn:
        resultado=conn.execute(text("""
            DELETE FROM periodo_fijo_discontinuo
            WHERE id_periodo=:id_periodo
        """),{
            "id_periodo":int(datos["id_periodo"])
        })
        print("FILAS ELIMINADAS:",resultado.rowcount)
    return jsonify(ok=True)


# ============================================================
# CALENDARIO
# ============================================================

# ---- Vista del calendario anual ----
@app.route("/calendario")
def calendario():
    ID_CENTRO=1
    anio=request.args.get("anio",type=int)
    anios_df=pd.read_sql("""
        SELECT DISTINCT YEAR(fecha) AS anio
        FROM calendario
        WHERE id_centro=%s
        ORDER BY anio DESC
    """,engine,params=(ID_CENTRO,))
    anios=anios_df["anio"].tolist()
    if not anios:
        return render_template("calendario.html",meses={},anios=[],anio=None)
    if anio not in anios:
        anio=anios[0]
    df=pd.read_sql("""
        SELECT *
        FROM calendario
        WHERE id_centro=%s
        AND YEAR(fecha)=%s
        ORDER BY fecha
    """,engine,params=(ID_CENTRO,anio))
    meses={}
    nombres_meses={
        1:"Enero",2:"Febrero",3:"Marzo",4:"Abril",
        5:"Mayo",6:"Junio",7:"Julio",8:"Agosto",
        9:"Septiembre",10:"Octubre",11:"Noviembre",12:"Diciembre"
    }
    for fila in df.to_dict(orient="records"):
        mes=nombres_meses[fila["fecha"].month]
        meses.setdefault(mes,[]).append(fila)
    for dias in meses.values():
        if dias:
            primer_dia=dias[0]["fecha"].weekday()
            for _ in range(primer_dia):
                dias.insert(0,None)
    return render_template(
        "calendario.html",
        meses=meses,
        anios=anios,
        anio=anio
    )


# ---- Actualizar un día del calendario ----
@app.route("/actualizar_dia",methods=["POST"])
def actualizar_dia():
    datos=request.get_json()
    with engine.begin() as conn:
        conn.execute(
            text("""
                UPDATE calendario
                SET
                    hora_apertura=:hora_apertura,
                    hora_cierre=:hora_cierre,
                    centro_abierto=:centro_abierto,
                    modificado=1
                WHERE fecha=:fecha
                AND id_centro=1
            """),
            {
                "hora_apertura":datos["hora_apertura"] or None,
                "hora_cierre":datos["hora_cierre"] or None,
                "centro_abierto":datos["abierto"],
                "fecha":datos["fecha"]
            }
        )
    return jsonify({"ok":True})


# ---- Importar calendario desde Excel ----
@app.route("/importar_calendario",methods=["POST"])
def importar_calendario():
    ID_CENTRO=1
    if "archivo" not in request.files:
        return jsonify(error="No se ha seleccionado ningún archivo."),400
    archivo=request.files["archivo"]
    if archivo.filename=="":
        return jsonify(error="No se ha seleccionado ningún archivo."),400
    if not archivo.filename.lower().endswith(".xlsx"):
        return jsonify(error="El archivo debe estar en formato .xlsx."),400
    try:
        df=pd.read_excel(archivo)
        columnas_obligatorias=[
            "fecha",
            "centro_abierto",
            "es_festivo",
            "dia_posterior_festivo",
            "hora_apertura",
            "hora_cierre"
        ]
        columnas_faltantes=[
            columna for columna in columnas_obligatorias
            if columna not in df.columns
        ]
        if columnas_faltantes:
            return jsonify(
                error="Faltan columnas obligatorias: "+", ".join(columnas_faltantes)
            ),400
        df=df[columnas_obligatorias].copy()
        df["fecha"]=pd.to_datetime(df["fecha"],errors="coerce")
        if df["fecha"].isna().any():
            return jsonify(error="Hay fechas no válidas en el Excel."),400
        if df["fecha"].duplicated().any():
            return jsonify(error="Hay fechas duplicadas en el Excel."),400
        def convertir_booleano(valor):
            if pd.isna(valor):
                return None
            if isinstance(valor,bool):
                return valor
            valor=str(valor).strip().upper()
            if valor in ["VERDADERO","TRUE","1"]:
                return True
            if valor in ["FALSO","FALSE","0"]:
                return False
            return None
        df["centro_abierto"]=df["centro_abierto"].apply(convertir_booleano)
        df["es_festivo"]=df["es_festivo"].apply(convertir_booleano)
        df["dia_posterior_festivo"]=df["dia_posterior_festivo"].apply(convertir_booleano)
        if df[[
            "centro_abierto",
            "es_festivo",
            "dia_posterior_festivo"
        ]].isna().any().any():
            return jsonify(
                error="Los campos 0/1 contienen valores no válidos."
            ),400
        def convertir_hora(valor):
            if pd.isna(valor):
                return None
            if hasattr(valor,"strftime"):
                return valor.strftime("%H:%M")
            valor=str(valor).strip()
            if valor=="" or valor.lower()=="nan":
                return None
            return valor[:5]
        df["hora_apertura"]=df["hora_apertura"].apply(convertir_hora)
        df["hora_cierre"]=df["hora_cierre"].apply(convertir_hora)
        actualizados=0
        insertados=0
        with engine.begin() as conn:
            for _,fila in df.iterrows():
                fecha=fila["fecha"].date()
                hora_apertura=fila["hora_apertura"]
                hora_cierre=fila["hora_cierre"]
                if pd.isna(hora_apertura):
                    hora_apertura=None
                if pd.isna(hora_cierre):
                    hora_cierre=None
                resultado=conn.execute(
                    text("""
                        UPDATE calendario
                        SET
                            centro_abierto=:centro_abierto,
                            es_festivo=:es_festivo,
                            dia_posterior_festivo=:dia_posterior_festivo,
                            hora_apertura=:hora_apertura,
                            hora_cierre=:hora_cierre
                        WHERE fecha=:fecha
                        AND id_centro=:id_centro
                    """),
                    {
                        "fecha":fecha,
                        "centro_abierto":bool(fila["centro_abierto"]),
                        "es_festivo":bool(fila["es_festivo"]),
                        "dia_posterior_festivo":bool(fila["dia_posterior_festivo"]),
                        "hora_apertura":hora_apertura,
                        "hora_cierre":hora_cierre,
                        "id_centro":ID_CENTRO
                    }
                )
                if resultado.rowcount>0:
                    actualizados+=1
                else:
                    conn.execute(
                        text("""
                            INSERT INTO calendario(
                                fecha,
                                centro_abierto,
                                es_festivo,
                                dia_posterior_festivo,
                                hora_apertura,
                                hora_cierre,
                                id_centro
                            )
                            VALUES(
                                :fecha,
                                :centro_abierto,
                                :es_festivo,
                                :dia_posterior_festivo,
                                :hora_apertura,
                                :hora_cierre,
                                :id_centro
                            )
                        """),
                        {
                            "fecha":fecha,
                            "centro_abierto":bool(fila["centro_abierto"]),
                            "es_festivo":bool(fila["es_festivo"]),
                            "dia_posterior_festivo":bool(fila["dia_posterior_festivo"]),
                            "hora_apertura":hora_apertura,
                            "hora_cierre":hora_cierre,
                            "id_centro":ID_CENTRO
                        }
                    )
                    insertados+=1
        return jsonify(
            ok=True,
            mensaje=(
                f"Calendario importado correctamente. "
                f"Se han añadido {insertados} días y "
                f"actualizado {actualizados} días."
            )
        )
    except Exception as e:
        print("ERROR AL IMPORTAR CALENDARIO:",e)
        return jsonify(
            error="No se ha podido importar el calendario."
        ),500


# ---- Exportar calendario a Excel ----
@app.route("/exportar_calendario")
def exportar_calendario():
    ID_CENTRO=1
    try:
        df=pd.read_sql("""
            SELECT
                fecha,
                centro_abierto,
                es_festivo,
                dia_posterior_festivo,
                hora_apertura,
                hora_cierre
            FROM calendario
            WHERE id_centro=%s
            ORDER BY fecha
        """,engine,params=(ID_CENTRO,))
        df["centro_abierto"]=df["centro_abierto"].astype(int)
        df["es_festivo"]=df["es_festivo"].astype(int)
        df["dia_posterior_festivo"]=df["dia_posterior_festivo"].astype(int)
        memoria=BytesIO()
        with pd.ExcelWriter(memoria,engine="openpyxl") as writer:
            df.to_excel(writer,index=False,sheet_name="Calendario")
            hoja=writer.sheets["Calendario"]
            hoja.column_dimensions["A"].width=14
            hoja.column_dimensions["B"].width=18
            hoja.column_dimensions["C"].width=15
            hoja.column_dimensions["D"].width=25
            hoja.column_dimensions["E"].width=18
            hoja.column_dimensions["F"].width=18
            for celda in hoja["A"][1:]:
                celda.number_format="dd/mm/yyyy"
            for fila in range(2,hoja.max_row+1):
                hoja.cell(fila,5).number_format="HH:MM"
                hoja.cell(fila,6).number_format="HH:MM"
        memoria.seek(0)
        return send_file(
            memoria,
            as_attachment=True,
            download_name="calendario.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        print("ERROR AL EXPORTAR CALENDARIO:",e)
        return jsonify(error="No se ha podido exportar el calendario."),500


# ============================================================
# GESTIÓN DE DATOS — DISPONIBILIDAD
# ============================================================

# ---- Periodos de disponibilidad de un trabajador (AJAX) ----
@app.route("/periodos_disponibilidad/<int:id_trabajador>")
def periodos_disponibilidad(id_trabajador):
    trabajador=pd.read_sql("""
        SELECT
            CONCAT(nombre,' ',apellidos) AS trabajador
        FROM trabajador
        WHERE id_trabajador=%s
    """,engine,params=(id_trabajador,))
    periodos=pd.read_sql("""
        SELECT
            id_disponibilidad,
            motivo,
            turno,
            fecha_inicio,
            fecha_fin
        FROM disponibilidad
        WHERE id_trabajador=%s
        ORDER BY fecha_inicio
    """,engine,params=(id_trabajador,))
    return jsonify({
        "trabajador":trabajador.iloc[0]["trabajador"],
        "periodos":periodos.assign(
            fecha_inicio=periodos["fecha_inicio"].astype(str),
            fecha_fin=periodos["fecha_fin"].astype(str)
        ).to_dict(orient="records")
    })


# ---- Nueva restricción de disponibilidad ----
@app.route("/nueva_disponibilidad", methods=["POST"])
def nueva_disponibilidad():

    datos = request.get_json()

    with engine.begin() as conn:

        conn.execute(text("""

            INSERT INTO disponibilidad(

                id_trabajador,
                motivo,
                turno,
                fecha_inicio,
                fecha_fin

            )
            VALUES(

                :id_trabajador,
                :motivo,
                :turno,
                :fecha_inicio,
                :fecha_fin

            )

        """), {

            "id_trabajador": datos["id_trabajador"],
            "fecha_inicio": datos["fecha_inicio"],
            "fecha_fin": datos["fecha_fin"],
            "motivo": datos["motivo"],
            "turno": datos["turno"]

        })

    return jsonify(ok=True)


# ---- Editar restricción de disponibilidad ----
@app.route("/editar_disponibilidad", methods=["POST"])
def editar_disponibilidad():

    datos = request.get_json()

    with engine.begin() as conn:

        conn.execute(text("""

            UPDATE disponibilidad

            SET

                id_trabajador=:id_trabajador,
                motivo=:motivo,
                turno=:turno,
                fecha_inicio=:fecha_inicio,
                fecha_fin=:fecha_fin

            WHERE id_disponibilidad=:id

        """),{

            "id": datos["id"],
            "id_trabajador": datos["id_trabajador"],
            "fecha_inicio": datos["fecha_inicio"],
            "fecha_fin": datos["fecha_fin"],
            "motivo": datos["motivo"],
            "turno": datos["turno"],

        })

    return jsonify(ok=True)


# ---- Eliminar restricción de disponibilidad ----
@app.route("/eliminar_disponibilidad", methods=["POST"])
def eliminar_disponibilidad():

    datos = request.get_json()

    with engine.begin() as conn:

        conn.execute(text("""

            DELETE FROM disponibilidad

            WHERE id_disponibilidad=:id

        """),{

            "id": datos["id"]

        })

    return jsonify(ok=True)


# ============================================================
# GESTIÓN DE DATOS — CENTRO Y TAREAS
# ============================================================

# ---- Vista del centro ----
@app.route("/centro")
def centro():
    return render_template("centro.html")


# ---- Nueva tarea ----
@app.route("/nueva_tarea", methods=["POST"])
def nueva_tarea():

    datos = request.get_json()

    with engine.begin() as conn:

        resultado = conn.execute(text("""

            INSERT INTO tarea(

                nombre,
                descripcion,
                activa

            )

            VALUES(

                :nombre,
                :descripcion,
                :activa

            )

        """), datos)

        id_tarea = resultado.lastrowid

        for id_trabajador in datos["trabajadores"]:

            conn.execute(text("""

                INSERT INTO trabajador_tarea(

                    id_trabajador,
                    id_tarea

                )

                VALUES(

                    :id_trabajador,
                    :id_tarea

                )

            """),{

                "id_trabajador": id_trabajador,
                "id_tarea": id_tarea

            })

    return jsonify(ok=True)


# ---- Editar tarea ----
@app.route("/editar_tarea", methods=["POST"])
def editar_tarea():
    datos=request.get_json()
    with engine.begin() as conn:
        conn.execute(text("""
            UPDATE tarea
            SET
                nombre=:nombre,
                descripcion=:descripcion,
                activa=:activa
            WHERE id_tarea=:id
        """),datos)
        conn.execute(text("""
            DELETE FROM trabajador_tarea
            WHERE id_tarea=:id
        """),{"id":datos["id"]})
        for id_trabajador in datos["trabajadores"]:
            conn.execute(text("""
                INSERT INTO trabajador_tarea(
                    id_trabajador,
                    id_tarea
                )
                VALUES(
                    :id_trabajador,
                    :id_tarea
                )
            """),{
                "id_trabajador":id_trabajador,
                "id_tarea":datos["id"]
            })
    return jsonify(ok=True)


# ---- Trabajadores asignados a una tarea (AJAX) ----
@app.route("/trabajadores_tarea/<int:id_tarea>")
def trabajadores_tarea(id_tarea):

    trabajadores=pd.read_sql("""

        SELECT
            t.id_trabajador,
            t.nombre,
            t.apellidos

        FROM trabajador_tarea tt

        JOIN trabajador t
            ON t.id_trabajador=tt.id_trabajador

        WHERE tt.id_tarea=%s

        ORDER BY t.apellidos,t.nombre

    """,engine,params=(id_tarea,))

    return trabajadores.to_json(
        orient="records",
        force_ascii=False
    )


# ---- Cambiar estado de una tarea ----
@app.route("/cambiar_estado_tarea", methods=["POST"])
def cambiar_estado_tarea():

    datos=request.get_json()

    with engine.begin() as conn:

        conn.execute(text("""
            UPDATE tarea
            SET activa=NOT activa
            WHERE id_tarea=:id
        """),{
            "id":datos["id"]
        })

    return jsonify(ok=True)


# ---- Eliminar tarea ----
@app.route("/eliminar_tarea", methods=["POST"])
def eliminar_tarea():
    datos=request.get_json()
    with engine.begin() as conn:
        conn.execute(text("""
            DELETE FROM trabajador_tarea
            WHERE id_tarea=:id
        """),{
            "id":datos["id"]
        })
        conn.execute(text("""
            DELETE FROM tarea
            WHERE id_tarea=:id
        """),{
            "id":datos["id"]
        })
    return jsonify(ok=True)


# ============================================================
# PROMOCIONES (API)
# ============================================================

# ---- Listado de promociones ----
@app.route("/api/promociones")
def api_promociones():

    try:

        df = pd.read_sql("""
            SELECT
                id_promocion,
                nombre,
                fecha_inicio,
                fecha_fin,
                tipo,
                descripcion,
                id_centro,
                COALESCE(tier, 3) AS tier
            FROM promocion
            WHERE id_centro = 1
            ORDER BY fecha_inicio
        """, engine)

        datos = df.to_dict(orient="records")

        for fila in datos:

            if pd.isna(fila.get("tipo")):
                fila["tipo"] = None

            if pd.isna(fila.get("descripcion")):
                fila["descripcion"] = None

            if pd.notna(fila.get("fecha_inicio")):
                fila["fecha_inicio"] = fila["fecha_inicio"].strftime("%Y-%m-%d")

            if pd.notna(fila.get("fecha_fin")):
                fila["fecha_fin"] = fila["fecha_fin"].strftime("%Y-%m-%d")

        return jsonify(datos)

    except Exception as e:

        print("ERROR PROMOCIONES:", e)

        return jsonify(
            error="No se han podido cargar las promociones."
        ), 500


# ---- Obtener una promoción ----
@app.route("/api/promociones/<int:id_promocion>", methods=["GET"])
def obtener_promocion(id_promocion):

    try:

        with engine.begin() as conn:

            resultado = conn.execute(
                text("""
                    SELECT
                        id_promocion,
                        nombre,
                        fecha_inicio,
                        fecha_fin,
                        tipo,
                        descripcion,
                        id_centro,
                        COALESCE(tier, 3) AS tier
                    FROM promocion
                    WHERE id_promocion = :id
                    AND id_centro = 1
                """),
                {
                    "id": id_promocion
                }
            )

            fila = resultado.mappings().first()

        if fila is None:

            return jsonify(
                error="No se ha encontrado la promoción."
            ), 404

        datos = dict(fila)

        # Convertir fechas a texto
        if datos.get("fecha_inicio") is not None:
            datos["fecha_inicio"] = datos["fecha_inicio"].strftime(
                "%Y-%m-%d"
            )

        if datos.get("fecha_fin") is not None:
            datos["fecha_fin"] = datos["fecha_fin"].strftime(
                "%Y-%m-%d"
            )

        # Evitar problemas con valores NULL
        if datos.get("tipo") is None:
            datos["tipo"] = ""

        if datos.get("descripcion") is None:
            datos["descripcion"] = ""

        return jsonify(datos)

    except Exception as e:

        print("ERROR OBTENIENDO PROMOCIÓN:", e)

        return jsonify(
            error="No se ha podido cargar la promoción."
        ), 500


# ---- Crear promoción ----
@app.route("/api/promociones",methods=["POST"])
def crear_promocion():
    datos=request.get_json()

    try:
        with engine.begin() as conn:
            conn.execute(
                text("""
                    INSERT INTO promocion(
                        nombre,
                        fecha_inicio,
                        fecha_fin,
                        tipo,
                        descripcion,
                        id_centro,
                        tier
                    )
                    VALUES(
                        :nombre,
                        :fecha_inicio,
                        :fecha_fin,
                        :tipo,
                        :descripcion,
                        1,
                        :tier
                    )
                """),
                {
                    "nombre":datos["nombre"],
                    "fecha_inicio":datos["fecha_inicio"],
                    "fecha_fin":datos["fecha_fin"],
                    "tipo":datos.get("tipo"),
                    "descripcion":datos.get("descripcion"),
                    "tier": int(datos.get("tier") or 3)
                }
            )

        return jsonify(ok=True)

    except Exception as e:
        print("ERROR CREANDO PROMOCIÓN:",e)
        return jsonify(
            error="No se ha podido guardar la promoción."
        ),500


# ---- Actualizar promoción ----
@app.route("/api/promociones/<int:id_promocion>",methods=["PUT"])
def actualizar_promocion(id_promocion):
    datos=request.get_json()

    try:
        with engine.begin() as conn:
            resultado=conn.execute(
                text("""
                    UPDATE promocion
                    SET
                        nombre=:nombre,
                        fecha_inicio=:fecha_inicio,
                        fecha_fin=:fecha_fin,
                        tipo=:tipo,
                        descripcion=:descripcion,
                        tier=:tier
                    WHERE id_promocion=:id
                    AND id_centro=1
                """),
                {
                    "id":id_promocion,
                    "nombre":datos["nombre"],
                    "fecha_inicio":datos["fecha_inicio"],
                    "fecha_fin":datos["fecha_fin"],
                    "tipo":datos.get("tipo"),
                    "descripcion":datos.get("descripcion"),
                    "tier": int(datos.get("tier") or 3)
                }
            )

        if resultado.rowcount==0:
            return jsonify(
                error="No se ha encontrado la promoción."
            ),404

        return jsonify(ok=True)

    except Exception as e:
        print("ERROR ACTUALIZANDO PROMOCIÓN:",e)
        return jsonify(
            error="No se ha podido actualizar la promoción."
        ),500


# ---- Eliminar promoción ----
@app.route("/api/promociones/<int:id_promocion>",methods=["DELETE"])
def eliminar_promocion(id_promocion):
    try:
        with engine.begin() as conn:
            resultado=conn.execute(
                text("""
                    DELETE FROM promocion
                    WHERE id_promocion=:id
                    AND id_centro=1
                """),
                {
                    "id":id_promocion
                }
            )

        if resultado.rowcount==0:
            return jsonify(
                error="No se ha encontrado la promoción."
            ),404

        return jsonify(ok=True)

    except Exception as e:
        print("ERROR ELIMINANDO PROMOCIÓN:",e)
        return jsonify(
            error="No se ha podido eliminar la promoción."
        ),500


# ---- Importar promociones desde Excel ----
@app.route("/importar_promociones",methods=["POST"])
def importar_promociones():

    ID_CENTRO=1

    if "archivo" not in request.files:
        return jsonify(
            error="No se ha seleccionado ningún archivo."
        ),400

    archivo=request.files["archivo"]

    if archivo.filename=="":
        return jsonify(
            error="No se ha seleccionado ningún archivo."
        ),400

    if not archivo.filename.lower().endswith(".xlsx"):
        return jsonify(
            error="El archivo debe estar en formato .xlsx."
        ),400

    try:

        df=pd.read_excel(archivo)

        columnas_obligatorias=[
            "nombre",
            "fecha_inicio",
            "fecha_fin",
            "tipo",
            "descripcion"
        ]

        columnas_faltantes=[
            columna
            for columna in columnas_obligatorias
            if columna not in df.columns
        ]

        if columnas_faltantes:
            return jsonify(
                error=
                "Faltan columnas obligatorias: "
                +", ".join(columnas_faltantes)
            ),400

        df=df[columnas_obligatorias].copy()

        df["fecha_inicio"]=pd.to_datetime(
            df["fecha_inicio"],
            errors="coerce"
        )

        df["fecha_fin"]=pd.to_datetime(
            df["fecha_fin"],
            errors="coerce"
        )

        if df["fecha_inicio"].isna().any() or df["fecha_fin"].isna().any():
            return jsonify(
                error="Hay fechas no válidas en el Excel."
            ),400

        if df["nombre"].isna().any():
            return jsonify(
                error="Hay promociones sin nombre."
            ),400

        with engine.begin() as conn:

            for _,fila in df.iterrows():

                conn.execute(
                    text("""
                        INSERT INTO promocion(
                            nombre,
                            fecha_inicio,
                            fecha_fin,
                            tipo,
                            descripcion,
                            id_centro
                        )
                        VALUES(
                            :nombre,
                            :fecha_inicio,
                            :fecha_fin,
                            :tipo,
                            :descripcion,
                            :id_centro
                        )
                    """),
                    {
                        "nombre":str(fila["nombre"]).strip(),

                        "fecha_inicio":
                            fila["fecha_inicio"].date(),

                        "fecha_fin":
                            fila["fecha_fin"].date(),

                        "tipo":
                            None
                            if pd.isna(fila["tipo"])
                            else str(fila["tipo"]).strip(),

                        "descripcion":
                            None
                            if pd.isna(fila["descripcion"])
                            else str(fila["descripcion"]).strip(),

                        "id_centro":ID_CENTRO
                    }
                )

        return jsonify(
            ok=True,
            mensaje=
            f"Promociones importadas correctamente. "
            f"Se han cargado {len(df)} promociones."
        )

    except Exception as e:

        print("ERROR AL IMPORTAR PROMOCIONES:",e)

        return jsonify(
            error="No se han podido importar las promociones."
        ),500


# ---- Exportar promociones a Excel ----
@app.route("/exportar_promociones")
def exportar_promociones():

    try:

        df=pd.read_sql("""
            SELECT
                nombre,
                fecha_inicio,
                fecha_fin,
                tipo,
                descripcion
            FROM promocion
            WHERE id_centro=1
            ORDER BY fecha_inicio
        """,engine)

        output=BytesIO()

        with pd.ExcelWriter(
            output,
            engine="openpyxl",
            date_format="DD/MM/YYYY"
        ) as writer:

            df.to_excel(
                writer,
                index=False,
                sheet_name="Promociones"
            )

        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="promociones.xlsx",
            mimetype=
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:

        print("ERROR EXPORTANDO PROMOCIONES:",e)

        return jsonify(
            error="No se han podido exportar las promociones."
        ),500


# ============================================================
# DATOS HISTÓRICOS — PEDIDOS
# ============================================================

# ---- Vista de datos históricos ----
@app.route("/historicos")
def historicos():
    return render_template("historicos.html")


# ---- Listado de pedidos históricos ----
@app.route("/api/pedidos_historicos")
def api_pedidos_historicos():

    try:

        df = pd.read_sql("""
            SELECT *
            FROM pedidohistorico
            WHERE id_centro = 1
            ORDER BY fecha
        """, engine)

        # Convertir todos los NaN / NaT a None
        # para que sean válidos en JSON
        df = df.astype(object).where(pd.notna(df), None)

        datos = df.to_dict(orient="records")

        for fila in datos:

            if fila.get("fecha") is not None:

                if hasattr(fila["fecha"], "strftime"):
                    fila["fecha"] = fila["fecha"].strftime("%Y-%m-%d")

        return jsonify(datos)

    except Exception as e:

        print("ERROR PEDIDOS HISTÓRICOS:", e)

        return jsonify(
            error="No se han podido cargar los pedidos históricos."
        ), 500


# ---- Obtener un día ----
@app.route("/api/pedidos_historicos/<int:id_pedido>",methods=["GET"])
def obtener_pedido_historico(id_pedido):
    try:
        df=pd.read_sql(
            """
            SELECT *
            FROM pedidohistorico
            WHERE id_pedido_historico=%s
            AND id_centro=1
            """,
            engine,
            params=(id_pedido,)
        )
        if df.empty:
            return jsonify(error="No se ha encontrado el registro."),404
        df=df.astype(object).where(pd.notna(df),None)
        fila=df.iloc[0].to_dict()
        if fila.get("fecha") is not None:
            if hasattr(fila["fecha"],"strftime"):
                fila["fecha"]=fila["fecha"].strftime("%Y-%m-%d")
        return jsonify(fila)
    except Exception as e:
        print("ERROR OBTENIENDO PEDIDO:",e)
        return jsonify(error="No se ha podido cargar el registro."),500


# ---- Crear día ----
@app.route("/api/pedidos_historicos",methods=["POST"])
def crear_pedido_historico():

    datos=request.get_json()

    try:
        mcia_pedidos=int(datos.get("envios_2h_mcia_general_pedidos",0))
        mcia_lineas=int(datos.get("envios_2h_mcia_general_lineas",0))
        food_pedidos=int(datos.get("envios_2h_food_pedidos",0))
        food_lineas=int(datos.get("envios_2h_food_lineas",0))
        encargos_pedidos=int(datos.get("encargos_pedidos",0))
        encargos_lineas=int(datos.get("encargos_lineas",0))
        home_pedidos=int(datos.get("home_delivery_pedidos",0))
        home_lineas=int(datos.get("home_delivery_lineas",0))
        devoluciones = datos.get("devoluciones")
        if devoluciones is not None:
            devoluciones = int(devoluciones)
        

        total_pedidos=(
            mcia_pedidos+
            food_pedidos+
            encargos_pedidos+
            home_pedidos
        )

        total_lineas=(
            mcia_lineas+
            food_lineas+
            encargos_lineas+
            home_lineas
        )

        with engine.begin() as conn:

            existe=conn.execute(
                text("""
                    SELECT COUNT(*)
                    FROM pedidohistorico
                    WHERE fecha=:fecha
                    AND id_centro=1
                """),
                {"fecha":datos["fecha"]}
            ).scalar()

            if existe:
                return jsonify(
                    error="Ya existe un registro para esa fecha."
                ),400

            conn.execute(
                text("""
                    INSERT INTO pedidohistorico(
                        fecha,
                        envios_2h_mcia_general_pedidos,
                        envios_2h_mcia_general_lineas,
                        envios_2h_food_pedidos,
                        envios_2h_food_lineas,
                        encargos_pedidos,
                        encargos_lineas,
                        home_delivery_pedidos,
                        home_delivery_lineas,
                        total_pedidos,
                        total_lineas,
                        devoluciones,
                        id_centro
                    )
                    VALUES(
                        :fecha,
                        :mcia_pedidos,
                        :mcia_lineas,
                        :food_pedidos,
                        :food_lineas,
                        :encargos_pedidos,
                        :encargos_lineas,
                        :home_pedidos,
                        :home_lineas,
                        :total_pedidos,
                        :total_lineas,
                        :devoluciones,
                        1
                    )
                """),
                {
                    "fecha":datos["fecha"],
                    "mcia_pedidos":mcia_pedidos,
                    "mcia_lineas":mcia_lineas,
                    "food_pedidos":food_pedidos,
                    "food_lineas":food_lineas,
                    "encargos_pedidos":encargos_pedidos,
                    "encargos_lineas":encargos_lineas,
                    "home_pedidos":home_pedidos,
                    "home_lineas":home_lineas,
                    "total_pedidos":total_pedidos,
                    "total_lineas":total_lineas,
                    "devoluciones": devoluciones
                }
            )

        return jsonify(ok=True)

    except Exception as e:
        print("ERROR CREANDO PEDIDO:",e)
        return jsonify(
            error="No se ha podido guardar el día."
        ),500


# ---- Editar día ----
@app.route("/api/pedidos_historicos/<int:id_pedido>",methods=["PUT"])
def actualizar_pedido_historico(id_pedido):

    datos=request.get_json()

    try:
        mcia_pedidos=int(datos.get("envios_2h_mcia_general_pedidos",0))
        mcia_lineas=int(datos.get("envios_2h_mcia_general_lineas",0))
        food_pedidos=int(datos.get("envios_2h_food_pedidos",0))
        food_lineas=int(datos.get("envios_2h_food_lineas",0))
        encargos_pedidos=int(datos.get("encargos_pedidos",0))
        encargos_lineas=int(datos.get("encargos_lineas",0))
        home_pedidos=int(datos.get("home_delivery_pedidos",0))
        home_lineas=int(datos.get("home_delivery_lineas",0))
        devoluciones = datos.get("devoluciones")
        if devoluciones is not None:
            devoluciones = int(devoluciones)

        total_pedidos=(
            mcia_pedidos+
            food_pedidos+
            encargos_pedidos+
            home_pedidos
        )

        total_lineas=(
            mcia_lineas+
            food_lineas+
            encargos_lineas+
            home_lineas
        )

        with engine.begin() as conn:

            resultado=conn.execute(
                text("""
                    UPDATE pedidohistorico
                    SET
                        fecha=:fecha,
                        envios_2h_mcia_general_pedidos=:mcia_pedidos,
                        envios_2h_mcia_general_lineas=:mcia_lineas,
                        envios_2h_food_pedidos=:food_pedidos,
                        envios_2h_food_lineas=:food_lineas,
                        encargos_pedidos=:encargos_pedidos,
                        encargos_lineas=:encargos_lineas,
                        home_delivery_pedidos=:home_pedidos,
                        home_delivery_lineas=:home_lineas,
                        total_pedidos=:total_pedidos,
                        devoluciones=:devoluciones,
                        total_lineas=:total_lineas
                    WHERE id_pedido_historico=:id
                    AND id_centro=1
                """),
                {
                    "id":id_pedido,
                    "fecha":datos["fecha"],
                    "mcia_pedidos":mcia_pedidos,
                    "mcia_lineas":mcia_lineas,
                    "food_pedidos":food_pedidos,
                    "food_lineas":food_lineas,
                    "encargos_pedidos":encargos_pedidos,
                    "encargos_lineas":encargos_lineas,
                    "home_pedidos":home_pedidos,
                    "home_lineas":home_lineas,
                    "total_pedidos":total_pedidos,
                    "total_lineas":total_lineas,
                    "devoluciones":devoluciones
                }
            )

        if resultado.rowcount==0:
            return jsonify(
                error="No se ha encontrado el registro."
            ),404

        return jsonify(ok=True)

    except Exception as e:
        print("ERROR ACTUALIZANDO PEDIDO:",e)
        return jsonify(
            error="No se ha podido actualizar el día."
        ),500


# ---- Eliminar día ----
@app.route("/api/pedidos_historicos/<int:id_pedido>",methods=["DELETE"])
def eliminar_pedido_historico(id_pedido):

    try:
        with engine.begin() as conn:

            resultado=conn.execute(
                text("""
                    DELETE FROM pedidohistorico
                    WHERE id_pedido_historico=:id
                    AND id_centro=1
                """),
                {"id":id_pedido}
            )

        if resultado.rowcount==0:
            return jsonify(
                error="No se ha encontrado el registro."
            ),404

        return jsonify(ok=True)

    except Exception as e:
        print("ERROR ELIMINANDO PEDIDO:",e)
        return jsonify(
            error="No se ha podido eliminar el día."
        ),500


# ---- Importar desde Excel ----
@app.route("/api/pedidos_historicos/importar",methods=["POST"])
def importar_pedidos_historicos():
    if "archivo" not in request.files:
        return jsonify(error="No se ha seleccionado ningún archivo."),400

    archivo=request.files["archivo"]

    if archivo.filename=="":
        return jsonify(error="No se ha seleccionado ningún archivo."),400

    if not archivo.filename.lower().endswith(".xlsx"):
        return jsonify(error="El archivo debe estar en formato .xlsx."),400

    try:
        df=pd.read_excel(archivo)

        columnas_excel=[
            "Fecha de venta",
            "Núm. pedidos 2h MCIA GENERAL",
            "Núm. Líneas 2h MCIA GENERAL",
            "Núm. pedidos 2h FOOD",
            "Núm. líneas 2h FOOD",
            "Núm. pedidos ENCARGOS",
            "Núm. líneas ENCARGOS",
            "Núm. pedidos HOME DELIVERY",
            "Núm. líneas HOME DELIVERY",
            "Devoluciones"
        ]

        faltantes=[
            columna
            for columna in columnas_excel
            if columna not in df.columns
        ]

        if faltantes:
            return jsonify(
                error="Faltan columnas: "+", ".join(faltantes)
            ),400

        df=df[columnas_excel].copy()

        df["Fecha de venta"]=pd.to_datetime(
            df["Fecha de venta"],
            errors="coerce"
        )

        if df["Fecha de venta"].isna().any():
            return jsonify(
                error="Hay fechas no válidas en el Excel."
            ),400

        if df["Fecha de venta"].duplicated().any():
            return jsonify(
                error="Hay fechas duplicadas en el Excel."
            ),400

        # Columnas de pedidos: no pueden estar vacías
        columnas_numericas=[
            "Núm. pedidos 2h MCIA GENERAL",
            "Núm. Líneas 2h MCIA GENERAL",
            "Núm. pedidos 2h FOOD",
            "Núm. líneas 2h FOOD",
            "Núm. pedidos ENCARGOS",
            "Núm. líneas ENCARGOS",
            "Núm. pedidos HOME DELIVERY",
            "Núm. líneas HOME DELIVERY"
        ]

        for columna in columnas_numericas:
            df[columna]=pd.to_numeric(
                df[columna],
                errors="coerce"
            )

            if df[columna].isna().any():
                return jsonify(
                    error=f"Hay valores no válidos en '{columna}'."
                ),400

            df[columna]=df[columna].astype(int)

        # Devoluciones puede estar vacío -> NULL
        df["Devoluciones"]=pd.to_numeric(
            df["Devoluciones"],
            errors="coerce"
        )

        with engine.begin() as conn:
            for _,fila in df.iterrows():

                fecha=fila["Fecha de venta"].date()

                mcia_pedidos=int(
                    fila["Núm. pedidos 2h MCIA GENERAL"]
                )

                mcia_lineas=int(
                    fila["Núm. Líneas 2h MCIA GENERAL"]
                )

                food_pedidos=int(
                    fila["Núm. pedidos 2h FOOD"]
                )

                food_lineas=int(
                    fila["Núm. líneas 2h FOOD"]
                )

                encargos_pedidos=int(
                    fila["Núm. pedidos ENCARGOS"]
                )

                encargos_lineas=int(
                    fila["Núm. líneas ENCARGOS"]
                )

                home_pedidos=int(
                    fila["Núm. pedidos HOME DELIVERY"]
                )

                home_lineas=int(
                    fila["Núm. líneas HOME DELIVERY"]
                )

                if pd.isna(fila["Devoluciones"]):
                    devoluciones=None
                else:
                    devoluciones=int(fila["Devoluciones"])

                total_pedidos=(
                    mcia_pedidos+
                    food_pedidos+
                    encargos_pedidos+
                    home_pedidos
                )

                total_lineas=(
                    mcia_lineas+
                    food_lineas+
                    encargos_lineas+
                    home_lineas
                )

                existe=conn.execute(
                    text("""
                        SELECT id_pedido_historico
                        FROM pedidohistorico
                        WHERE fecha=:fecha
                        AND id_centro=1
                    """),
                    {"fecha":fecha}
                ).fetchone()

                datos={
                    "fecha":fecha,
                    "mcia_pedidos":mcia_pedidos,
                    "mcia_lineas":mcia_lineas,
                    "food_pedidos":food_pedidos,
                    "food_lineas":food_lineas,
                    "encargos_pedidos":encargos_pedidos,
                    "encargos_lineas":encargos_lineas,
                    "home_pedidos":home_pedidos,
                    "home_lineas":home_lineas,
                    "total_pedidos":total_pedidos,
                    "total_lineas":total_lineas,
                    "devoluciones":devoluciones
                }

                if existe:
                    conn.execute(
                        text("""
                            UPDATE pedidohistorico
                            SET
                                envios_2h_mcia_general_pedidos=:mcia_pedidos,
                                envios_2h_mcia_general_lineas=:mcia_lineas,
                                envios_2h_food_pedidos=:food_pedidos,
                                envios_2h_food_lineas=:food_lineas,
                                encargos_pedidos=:encargos_pedidos,
                                encargos_lineas=:encargos_lineas,
                                home_delivery_pedidos=:home_pedidos,
                                home_delivery_lineas=:home_lineas,
                                total_pedidos=:total_pedidos,
                                total_lineas=:total_lineas,
                                devoluciones=:devoluciones
                            WHERE fecha=:fecha
                            AND id_centro=1
                        """),
                        datos
                    )

                else:
                    conn.execute(
                        text("""
                            INSERT INTO pedidohistorico(
                                fecha,
                                envios_2h_mcia_general_pedidos,
                                envios_2h_mcia_general_lineas,
                                envios_2h_food_pedidos,
                                envios_2h_food_lineas,
                                encargos_pedidos,
                                encargos_lineas,
                                home_delivery_pedidos,
                                home_delivery_lineas,
                                total_pedidos,
                                total_lineas,
                                devoluciones,
                                id_centro
                            )
                            VALUES(
                                :fecha,
                                :mcia_pedidos,
                                :mcia_lineas,
                                :food_pedidos,
                                :food_lineas,
                                :encargos_pedidos,
                                :encargos_lineas,
                                :home_pedidos,
                                :home_lineas,
                                :total_pedidos,
                                :total_lineas,
                                :devoluciones,
                                1
                            )
                        """),
                        datos
                    )

        return jsonify(
            ok=True,
            mensaje=f"Se han importado {len(df)} días correctamente."
        )

    except Exception as e:
        print("ERROR IMPORTANDO PEDIDOS:",e)

        return jsonify(
            error="No se han podido importar los pedidos históricos."
        ),500

# ---- Exportar a Excel ----
@app.route("/api/pedidos_historicos/exportar")
def exportar_pedidos_historicos():

    try:

        df = pd.read_sql("""
            SELECT
                fecha,
                envios_2h_mcia_general_pedidos,
                envios_2h_mcia_general_lineas,
                envios_2h_food_pedidos,
                envios_2h_food_lineas,
                encargos_pedidos,
                encargos_lineas,
                home_delivery_pedidos,
                home_delivery_lineas,
                total_pedidos,
                total_lineas,
                devoluciones
            FROM pedidohistorico
            WHERE id_centro = 1
            ORDER BY fecha
        """, engine)

        df.columns = [
            "Fecha de venta",
            "Núm. pedidos 2h MCIA GENERAL",
            "Núm. líneas 2h MCIA GENERAL",
            "Núm. pedidos 2h FOOD",
            "Núm. líneas 2h FOOD",
            "Núm. pedidos ENCARGOS",
            "Núm. líneas ENCARGOS",
            "Núm. pedidos HOME DELIVERY",
            "Núm. líneas HOME DELIVERY",
            "Total PEDIDOS",
            "Total LÍNEAS",
            "Devoluciones"
        ]

        output = BytesIO()

        with pd.ExcelWriter(
            output,
            engine="openpyxl",
            date_format="DD/MM/YYYY"
        ) as writer:

            df.to_excel(
                writer,
                index=False,
                sheet_name="Pedidos históricos"
            )

            hoja = writer.book["Pedidos históricos"]

            # Formato de fechas
            for celda in hoja["A"][1:]:
                celda.number_format = "DD/MM/YYYY"

            # Ancho de columnas
            anchos = {
                "A": 16,
                "B": 28,
                "C": 30,
                "D": 24,
                "E": 25,
                "F": 24,
                "G": 25,
                "H": 28,
                "I": 30,
                "J": 16,
                "K": 16,
                "L": 16
            }

            for columna, ancho in anchos.items():
                hoja.column_dimensions[columna].width = ancho

        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="pedidos_historicos.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:

        print("ERROR EXPORTANDO PEDIDOS HISTÓRICOS:", e)

        return jsonify(
            error="No se han podido exportar los pedidos históricos."
        ), 500

# ============================================================
# PREDICCIÓN
# ============================================================

# ---- Vista de predicción ----
@app.route("/prediccion")
def prediccion():
    return render_template("prediccion.html")


# ---- Obtener predicciones generadas ----
@app.route("/api/prediccion")
def api_prediccion():
    try:
        fecha_inicio = request.args.get("fecha_inicio")
        fecha_fin = request.args.get("fecha_fin")

        query = """
            SELECT
                id_prediccion,
                fecha,
                dia_semana,
                pedidos_previstos,
                pedidos_acumulados,
                horas_necesarias,
                limite_inferior,
                limite_superior,
                fecha_generacion,
                id_centro
            FROM prediccion
            WHERE id_centro = 1
        """
        params = {}

        if fecha_inicio:
            query += " AND fecha >= :fecha_inicio"
            params["fecha_inicio"] = fecha_inicio

        if fecha_fin:
            query += " AND fecha <= :fecha_fin"
            params["fecha_fin"] = fecha_fin

        query += " ORDER BY fecha"

        df = pd.read_sql(text(query), engine, params=params)

        df = df.astype(object).where(pd.notna(df), None)

        datos = df.to_dict(orient="records")

        for fila in datos:
            if fila.get("fecha") is not None:
                fila["fecha"] = fila["fecha"].strftime("%Y-%m-%d")

            if fila.get("fecha_generacion") is not None:
                fila["fecha_generacion"] = fila["fecha_generacion"].strftime("%Y-%m-%d %H:%M:%S")

        return jsonify(datos)

    except Exception as e:
        print("ERROR PREDICCIÓN:", e)
        return jsonify({
            "error": "No se han podido cargar las predicciones."
        }), 500


@app.route("/api/periodos")
def api_periodos():
    try:
        predicciones = pd.read_sql(text("""
            SELECT DATE_FORMAT(fecha, '%Y-%m') AS periodo,
                   MIN(fecha) AS fecha_inicio, MAX(fecha) AS fecha_fin,
                   MAX(fecha_generacion) AS fecha_generacion
            FROM prediccion
            WHERE id_centro = 1
            GROUP BY DATE_FORMAT(fecha, '%Y-%m')
            ORDER BY periodo
        """), engine).to_dict(orient="records")

        with engine.begin() as conn:
            planificados = conn.execute(text("""
                SELECT DISTINCT DATE_FORMAT(fecha, '%Y-%m') AS periodo
                FROM calendarizacion
                WHERE id_version = (
                    SELECT id_version FROM planificacion_version
                    WHERE activa = 1 ORDER BY id_version DESC LIMIT 1
                )
            """)).scalars().all()

        planificados = set(planificados)
        for periodo in predicciones:
            periodo["planificada"] = periodo["periodo"] in planificados
            for clave in ("fecha_inicio", "fecha_fin"):
                if periodo[clave] is not None:
                    periodo[clave] = periodo[clave].strftime("%Y-%m-%d")
            if periodo["fecha_generacion"] is not None:
                periodo["fecha_generacion"] = periodo["fecha_generacion"].strftime("%Y-%m-%d %H:%M:%S")

        return jsonify(predicciones)
    except Exception as e:
        print("ERROR PERIODOS:", e)
        return jsonify({"error": "No se han podido cargar los periodos."}), 500


@app.route("/api/periodos/<periodo>", methods=["DELETE"])
def eliminar_periodo(periodo):
    try:
        inicio = pd.to_datetime(f"{periodo}-01").date()
        fin = (pd.Timestamp(inicio) + pd.offsets.MonthEnd(1)).date()
    except Exception:
        return jsonify({"error": "El periodo no tiene un formato válido."}), 400

    try:
        with engine.begin() as conn:
            conn.execute(text("""
                DELETE FROM prediccion
                WHERE id_centro = 1 AND fecha BETWEEN :inicio AND :fin
            """), {"inicio": inicio, "fin": fin})
            conn.execute(text("""
                DELETE FROM calendarizacion
                                WHERE fecha BETWEEN :inicio AND :fin
            """), {"inicio": inicio, "fin": fin})

        return jsonify({"ok": True, "mensaje": "Periodo eliminado correctamente."})
    except Exception as e:
        print("ERROR ELIMINANDO PERIODO:", e)
        return jsonify({"error": "No se ha podido eliminar el periodo."}), 500


# ---- Generar nueva predicción ----
@app.route("/api/prediccion/generar", methods=["POST"])
def generar_prediccion():
    try:
        import subprocess
        import sys
        import os

        # ====================================================
        # RECIBIR FECHAS DESDE LA INTERFAZ
        # ====================================================

        datos = request.get_json()

        if not datos:
            return jsonify({
                "error": "No se han recibido las fechas."
            }), 400

        fecha_inicio = datos.get("fecha_inicio")
        fecha_fin = datos.get("fecha_fin")

        if not fecha_inicio or not fecha_fin:
            return jsonify({
                "error": "Debes seleccionar una fecha de inicio y una fecha de fin."
            }), 400

        # Comprobar que el periodo es válido

        from datetime import datetime

        try:
            inicio = datetime.strptime(
                fecha_inicio,
                "%Y-%m-%d"
            ).date()

            fin = datetime.strptime(
                fecha_fin,
                "%Y-%m-%d"
            ).date()

        except ValueError:
            return jsonify({
                "error": "El formato de las fechas no es válido."
            }), 400

        if inicio > fin:
            return jsonify({
                "error": "La fecha de inicio no puede ser posterior a la fecha de fin."
            }), 400

        


        # ====================================================
        # EJECUTAR MODELO
        # ====================================================

        script = os.path.join(
            os.path.dirname(__file__),
            "..",
            "..",
            "FASE_1_modelo",
            "entrenamiento_prophet.py"
        )

        resultado = subprocess.run(
            [
                sys.executable,
                script,
                "--fecha-inicio",
                fecha_inicio,
                "--fecha-fin",
                fecha_fin
            ],
            capture_output=True,
            text=True
        )


        # ====================================================
        # COMPROBAR RESULTADO
        # ====================================================

        if resultado.returncode != 0:

            print("ERROR ENTRENAMIENTO:")
            print(resultado.stderr)

            return jsonify({
                "error": "No se ha podido generar la predicción.",
                "detalle": resultado.stderr
            }), 500


        print("PREDICCIÓN GENERADA:")
        print(resultado.stdout)


        return jsonify({
            "ok": True,
            "mensaje": "Predicción generada correctamente.",
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin
        })


    except Exception as e:

        print("ERROR:", e)

        return jsonify({
            "error": str(e)
        }), 500


    
# ============================================================
# PLANIFICACIÓN
# ============================================================

@app.route("/planificacion")
def planificacion():
    return render_template("planificacion.html")

# ============================================================
# PLANIFICACIÓN — DATOS DE CALENDARIZACIÓN
# ============================================================

@app.route("/api/planificacion")
def api_planificacion():

    try:
        fecha_inicio = request.args.get("fecha_inicio")
        fecha_fin = request.args.get("fecha_fin")

        query = """
            SELECT
                c.fecha,
                c.num_semana,
                c.id_tarea,
                t.nombre AS tarea,
                c.id_trabajador,
                tr.nombre AS trabajador,
                tr.apellidos AS apellidos,
                tr.fijo_discontinuo,
                c.turno

            FROM calendarizacion c

            INNER JOIN tarea t
                ON c.id_tarea = t.id_tarea

            INNER JOIN trabajador tr
                ON c.id_trabajador = tr.id_trabajador

            WHERE tr.activo = 1
              AND c.id_version = (
                  SELECT id_version
                  FROM planificacion_version
                  WHERE activa = 1
                  ORDER BY id_version DESC
                  LIMIT 1
              )
        """
        params = {}

        if fecha_inicio:
            query += " AND c.fecha >= :fecha_inicio"
            params["fecha_inicio"] = fecha_inicio

        if fecha_fin:
            query += " AND c.fecha <= :fecha_fin"
            params["fecha_fin"] = fecha_fin

        query += """
            ORDER BY
                c.fecha,
                c.turno,
                t.nombre,
                tr.nombre
        """

        df = pd.read_sql(text(query), engine, params=params)

        df = df.astype(object).where(pd.notna(df), None)

        datos = df.to_dict(orient="records")

        for fila in datos:

            if fila.get("fecha") is not None:
                fila["fecha"] = fila["fecha"].strftime("%Y-%m-%d")

            fila["fijo_discontinuo"] = bool(
                fila.get("fijo_discontinuo")
            )

        return jsonify(datos)

    except Exception as e:

        print("ERROR PLANIFICACIÓN:", e)

        return jsonify({
            "error": "No se ha podido cargar la planificación."
        }), 500


# ============================================================
# PLANIFICACIÓN — TAREAS
# ============================================================

@app.route("/api/tareas")
def api_tareas():

    try:

        df = pd.read_sql("""
            SELECT
                id_tarea,
                nombre
            FROM tarea
            WHERE activa = 1
            ORDER BY id_tarea
        """, engine)

        return jsonify(
            df.to_dict(orient="records")
        )

    except Exception as e:

        print("ERROR TAREAS:", e)

        return jsonify({
            "error": "No se han podido cargar las tareas."
        }), 500


# EXPORTAR PLANIFICACIÓN
@app.route("/api/planificacion/exportar")
def exportar_planificacion():

    try:

        query = """
            SELECT
                c.fecha,
                c.id_tarea,
                t.nombre AS tarea,
                c.id_trabajador,
                tr.nombre AS trabajador,
                c.turno
            FROM calendarizacion c
            JOIN tarea t
                ON c.id_tarea = t.id_tarea
            JOIN trabajador tr
                ON c.id_trabajador = tr.id_trabajador
            WHERE t.activa = 1
              AND c.id_version = (
                  SELECT id_version
                  FROM planificacion_version
                  WHERE activa = 1
                  ORDER BY id_version DESC
                  LIMIT 1
              )
            ORDER BY c.fecha, c.id_tarea, c.turno, c.id_trabajador
        """

        df = pd.read_sql(query, engine)

        if df.empty:
            return jsonify({
                "error": "No hay planificación para exportar."
            }), 404

        df["fecha"] = pd.to_datetime(df["fecha"]).dt.date

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        # Eliminamos la hoja que crea Excel automáticamente
        hoja_inicial = wb.active
        wb.remove(hoja_inicial)

        # ----------------------------------------------------
        # FECHAS
        # ----------------------------------------------------

        fechas = pd.date_range(
            start=df["fecha"].min(),
            end=df["fecha"].max(),
            freq="D"
        ).date

        # ----------------------------------------------------
        # TAREAS ACTIVAS
        # ----------------------------------------------------

        tareas = (
            df[["id_tarea", "tarea"]]
            .drop_duplicates()
            .sort_values("id_tarea")
        )

        # ----------------------------------------------------
        # ESTILOS
        # ----------------------------------------------------

        borde = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC")
        )

        def crear_hoja(nombre_hoja, turno):

            ws = wb.create_sheet(nombre_hoja)

            # ----------------------------------------------
            # CABECERA
            # ----------------------------------------------

            ws.cell(
                row=1,
                column=1,
                value="TAREA"
            )

            for columna, fecha in enumerate(fechas, start=2):

                celda = ws.cell(
                    row=1,
                    column=columna,
                    value=fecha
                )

                celda.number_format = "dd/mm/yyyy"

            # ----------------------------------------------
            # ESTILO CABECERA
            # ----------------------------------------------

            for celda in ws[1]:

                celda.font = Font(
                    bold=True,
                    color="FFFFFF"
                )

                celda.fill = PatternFill(
                    fill_type="solid",
                    fgColor="087F3F"
                )

                celda.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

                celda.border = borde

            ws.row_dimensions[1].height = 30

            # ----------------------------------------------
            # TAREAS
            # ----------------------------------------------

            fila = 2

            for _, tarea in tareas.iterrows():

                id_tarea = tarea["id_tarea"]

                ws.cell(
                    row=fila,
                    column=1,
                    value=tarea["tarea"]
                )

                # Estilo nombre tarea
                celda_tarea = ws.cell(
                    row=fila,
                    column=1
                )

                celda_tarea.font = Font(
                    bold=True
                )

                celda_tarea.fill = PatternFill(
                    fill_type="solid",
                    fgColor="E8FAF0"
                )

                celda_tarea.alignment = Alignment(
                    vertical="center",
                    wrap_text=True
                )

                celda_tarea.border = borde

                # ------------------------------------------
                # CADA DÍA
                # ------------------------------------------

                for columna, fecha in enumerate(
                    fechas,
                    start=2
                ):

                    registros = df[
                        (df["id_tarea"] == id_tarea)
                        &
                        (df["fecha"] == fecha)
                        &
                        (df["turno"] == turno)
                    ]

                    if registros.empty:

                        valor = "—"

                    else:

                        trabajadores = []

                        for _, trabajador in registros.iterrows():

                            trabajadores.append(
                                trabajador["trabajador"]
                            )

                        valor = "\n".join(
                            trabajadores
                        )

                    celda = ws.cell(
                        row=fila,
                        column=columna,
                        value=valor
                    )

                    celda.alignment = Alignment(
                        horizontal="center",
                        vertical="center",
                        wrap_text=True
                    )

                    celda.border = borde

                ws.row_dimensions[fila].height = 60

                fila += 1

            # ----------------------------------------------
            # ANCHOS
            # ----------------------------------------------

            ws.column_dimensions["A"].width = 35

            for columna in range(
                2,
                len(fechas) + 2
            ):

                letra = get_column_letter(
                    columna
                )

                ws.column_dimensions[
                    letra
                ].width = 18

            # ----------------------------------------------
            # CONGELAR
            # ----------------------------------------------

            ws.freeze_panes = "B2"

            # ----------------------------------------------
            # FILTRO
            # ----------------------------------------------

            ws.auto_filter.ref = (
                f"A1:"
                f"{get_column_letter(len(fechas) + 1)}"
                f"{fila - 1}"
            )

            return ws

        # ----------------------------------------------------
        # CREAR LAS DOS HOJAS
        # ----------------------------------------------------

        crear_hoja(
            "Mañanas",
            0
        )

        crear_hoja(
            "Tardes",
            1
        )

        # ----------------------------------------------------
        # GENERAR ARCHIVO
        # ----------------------------------------------------

        archivo = BytesIO()

        wb.save(archivo)

        archivo.seek(0)

        return send_file(
            archivo,
            as_attachment=True,
            download_name="planificacion.xlsx",
            mimetype=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            )
        )

    except Exception as e:

        print(
            "ERROR EXPORTANDO PLANIFICACIÓN:",
            e
        )

        return jsonify({
            "error": "No se ha podido generar el Excel."
        }), 500

# GENERAR PLANIFICACIÓN
@app.route("/api/planificacion/generar", methods=["POST"])
def generar_planificacion():

    try:

        datos = request.get_json(silent=True) or {}
        fecha_inicio = datos.get("fecha_inicio")
        fecha_fin = datos.get("fecha_fin")

        if fecha_inicio and fecha_fin:
            try:
                inicio_dt = pd.to_datetime(fecha_inicio).date()
                fin_dt = pd.to_datetime(fecha_fin).date()
                if (fin_dt - inicio_dt).days + 1 < 30:
                    return jsonify({
                        "ok": False,
                        "error": "Selecciona un periodo mínimo de 1 mes para generar la planificación."
                    }), 400
            except Exception:
                return jsonify({
                    "ok": False,
                    "error": "Las fechas seleccionadas no tienen un formato válido."
                }), 400

        with engine.begin() as conn:
            sql = "SELECT COUNT(*) FROM prediccion WHERE id_centro = 1"
            if fecha_inicio and fecha_fin:
                sql += " AND fecha BETWEEN :fecha_inicio AND :fecha_fin"
            total_prediccion = conn.execute(text(sql), {
                "fecha_inicio": fecha_inicio,
                "fecha_fin": fecha_fin,
            }).scalar()

        if total_prediccion is None or int(total_prediccion) == 0:
            return jsonify({
                "ok": False,
                "error": "No hay datos de predicción para el rango solicitado. Genera primero la predicción desde la pantalla de Predicción."
            }), 400

        ruta_script = (
            Path(__file__).resolve().parents[2]
            / "FASE_3_calendarizacion"
            / "calPRUEBA.py"
        )

        comando = [sys.executable, str(ruta_script)]
        if fecha_inicio:
            comando.extend(["--fecha-inicio", str(fecha_inicio)])
        if fecha_fin:
            comando.extend(["--fecha-fin", str(fecha_fin)])

        job_id = uuid.uuid4().hex
        with planificacion_trabajos_lock:
            if any(
                trabajo["estado"] == "ejecutando"
                for trabajo in planificacion_trabajos.values()
            ):
                return jsonify({
                    "ok": False,
                    "error": "Ya hay una planificación en curso. Espera a que termine."
                }), 409
            planificacion_trabajos[job_id] = {"estado": "ejecutando"}

        planificacion_executor.submit(
            ejecutar_generacion_planificacion,
            job_id,
            comando,
            ruta_script,
            fecha_inicio,
            fecha_fin
        )

        return jsonify({
            "ok": True,
            "en_curso": True,
            "job_id": job_id,
            "mensaje": "La planificación se está generando en segundo plano."
        }), 202

    except Exception as e:

        print("ERROR EJECUTANDO CALPRUEBA:", e)

        return jsonify({
            "ok": False,
            "error": str(e)
        }), 500


@app.route("/api/planificacion/generar/<job_id>")
def estado_generacion_planificacion(job_id):
    with planificacion_trabajos_lock:
        trabajo = planificacion_trabajos.get(job_id)

    if trabajo is None:
        return jsonify({"ok": False, "error": "No se ha encontrado el trabajo."}), 404

    if trabajo["estado"] == "ejecutando":
        return jsonify({"ok": True, "estado": "ejecutando"}), 202

    if trabajo["estado"] == "error":
        return jsonify({
            "ok": False,
            "estado": "error",
            "error": "No se ha podido generar la planificación.",
            "detalle": trabajo.get("detalle", "")
        }), 500

    return jsonify({
        "ok": True,
        "estado": "completado",
        "mensaje": "Planificación generada correctamente.",
        "filas_guardadas": trabajo.get("filas_guardadas"),
        "version_guardada": trabajo.get("version_guardada"),
        "salida": trabajo.get("salida", "")
    })

# ============================================================
# PLANIFICACIÓN - CALENDARIO INDIVIDUAL DEL TRABAJADOR
# ============================================================

@app.route("/calendarios-trabajadores")
def calendarios_trabajadores():
    return render_template("calendarios_trabajadores.html")

@app.route("/calendario-trabajador/<int:id_trabajador>")
def calendario_trabajador(id_trabajador):
    return render_template(
        "calendario_trabajador.html",
        id_trabajador=id_trabajador
    )

def convertir_valor(valor):

    if pd.isna(valor):
        return None

    if hasattr(valor, "item"):
        return valor.item()

    return valor

@app.route("/api/calendario-trabajador/<int:id_trabajador>")
def api_calendario_trabajador(id_trabajador):

    try:

        # ====================================================
        # CALENDARIZACIÓN
        # ====================================================

        df = pd.read_sql("""
            SELECT
                c.fecha,
                c.id_trabajador,
                c.id_tarea,
                c.turno,
                t.nombre AS tarea
            FROM calendarizacion c
            LEFT JOIN tarea t
                ON c.id_tarea = t.id_tarea
            WHERE c.id_trabajador = %s
              AND c.id_version = (
                  SELECT id_version
                  FROM planificacion_version
                  WHERE activa = 1
                  ORDER BY id_version DESC
                  LIMIT 1
              )
            ORDER BY c.fecha
        """, engine, params=(id_trabajador,))


        # ====================================================
        # TRABAJADOR
        # ====================================================

        trabajador = pd.read_sql("""
            SELECT
                tr.id_trabajador,
                tr.numero_vendedor,
                tr.nombre,
                tr.apellidos,
                tr.correo,
                tr.activo,
                tr.id_contrato,
                tr.id_centro,
                tr.estado,
                tr.disponibilidad,
                tr.fijo_discontinuo,

                co.nombre AS contrato,
                co.jornada,
                co.horas_por_turno

            FROM trabajador tr

            LEFT JOIN contrato co
                ON tr.id_contrato = co.id_contrato

            WHERE tr.id_trabajador = %s

            LIMIT 1
        """, engine, params=(id_trabajador,))


        if trabajador.empty:

            return jsonify({
                "error": "Trabajador no encontrado."
            }), 404


        # ====================================================
        # DISPONIBILIDAD / VACACIONES
        # ====================================================

        disponibilidad = pd.read_sql("""
            SELECT
                id_disponibilidad,
                id_trabajador,
                fecha_inicio,
                fecha_fin,
                motivo,
                turno,
                id_turno
            FROM disponibilidad
            WHERE id_trabajador = %s
            ORDER BY fecha_inicio
        """, engine, params=(id_trabajador,))


        # ====================================================
        # CALENDARIZACIÓN → JSON
        # ====================================================

        registros = []

        for _, fila in df.iterrows():

            registros.append({

                "fecha": (
                    fila["fecha"].strftime("%Y-%m-%d")
                    if pd.notna(fila["fecha"])
                    else None
                ),

                "id_trabajador": (
                    int(fila["id_trabajador"])
                    if pd.notna(fila["id_trabajador"])
                    else None
                ),

                "id_tarea": (
                    int(fila["id_tarea"])
                    if pd.notna(fila["id_tarea"])
                    else None
                ),

                "turno": convertir_valor(
                    fila["turno"]
                ),

                "tarea": (
                    str(fila["tarea"])
                    if pd.notna(fila["tarea"])
                    else None
                )
            })


        # ====================================================
        # DISPONIBILIDAD → JSON
        # ====================================================

        vacaciones = []

        for _, fila in disponibilidad.iterrows():

            vacaciones.append({

                "fecha_inicio": (
                    fila["fecha_inicio"].strftime("%Y-%m-%d")
                    if pd.notna(fila["fecha_inicio"])
                    else None
                ),

                "fecha_fin": (
                    fila["fecha_fin"].strftime("%Y-%m-%d")
                    if pd.notna(fila["fecha_fin"])
                    else None
                ),

                "motivo": (
                    str(fila["motivo"])
                    if pd.notna(fila["motivo"])
                    else None
                ),

                "turno": convertir_valor(
                    fila["turno"]
                )
            })


        # ====================================================
        # INFORMACIÓN DEL TRABAJADOR
        # ====================================================

        tr = trabajador.iloc[0]

        datos_trabajador = {

            "id_trabajador":
                int(tr["id_trabajador"])
                if pd.notna(tr["id_trabajador"])
                else None,

            "numero_vendedor":
                convertir_valor(tr["numero_vendedor"]),

            "nombre":
                str(tr["nombre"])
                if pd.notna(tr["nombre"])
                else None,

            "apellidos":
                str(tr["apellidos"])
                if pd.notna(tr["apellidos"])
                else None,

            "correo":
                str(tr["correo"])
                if pd.notna(tr["correo"])
                else None,

            "activo":
                bool(tr["activo"])
                if pd.notna(tr["activo"])
                else None,

            "id_contrato":
                int(tr["id_contrato"])
                if pd.notna(tr["id_contrato"])
                else None,

            "contrato":
                str(tr["contrato"])
                if pd.notna(tr["contrato"])
                else None,

            "jornada":
                float(tr["jornada"])
                if pd.notna(tr["jornada"])
                else None,

            "horas_por_turno":
                float(tr["horas_por_turno"])
                if pd.notna(tr["horas_por_turno"])
                else None,

            "id_centro":
                int(tr["id_centro"])
                if pd.notna(tr["id_centro"])
                else None,

            "estado":
                convertir_valor(tr["estado"]),

            "disponibilidad":
                convertir_valor(tr["disponibilidad"]),

            "fijo_discontinuo":
                bool(tr["fijo_discontinuo"])
                if pd.notna(tr["fijo_discontinuo"])
                else None
        }


        # ====================================================
        # RESPUESTA
        # ====================================================

        return jsonify({

            "trabajador":
                datos_trabajador,

            "calendarizacion":
                registros,

            "disponibilidad":
                vacaciones

        })


    except Exception as e:

        print(
            "ERROR CALENDARIO TRABAJADOR:"
        )

        print(e)

        return jsonify({

            "error":
                "No se ha podido cargar el calendario del trabajador."

        }), 500


# ============================================================
# LISTA DE TRABAJADORES
# ============================================================

@app.route("/api/calendarios-trabajadores")
def api_calendarios_trabajadores():

    try:

        df = pd.read_sql("""
            SELECT
                id_trabajador,
                nombre,
                apellidos,
                numero_vendedor,
                activo,
                fijo_discontinuo
            FROM trabajador
            WHERE id_centro = 1
              AND activo = 1
            ORDER BY apellidos, nombre
        """, engine)


        df = df.astype(object).where(
            pd.notna(df),
            None
        )


        # Convertir valores numpy a tipos Python

        datos = []

        for _, fila in df.iterrows():

            datos.append({

                "id_trabajador":
                    int(fila["id_trabajador"])
                    if fila["id_trabajador"] is not None
                    else None,

                "nombre":
                    str(fila["nombre"])
                    if fila["nombre"] is not None
                    else None,

                "apellidos":
                    str(fila["apellidos"])
                    if fila["apellidos"] is not None
                    else None,

                "numero_vendedor":
                    convertir_valor(
                        fila["numero_vendedor"]
                    ),

                "activo":
                    bool(fila["activo"])
                    if fila["activo"] is not None
                    else None,

                "fijo_discontinuo":
                    bool(fila["fijo_discontinuo"])
                    if fila["fijo_discontinuo"] is not None
                    else None
            })


        return jsonify(datos)


    except Exception as e:

        print(
            "ERROR TRABAJADORES:"
        )

        print(e)

        return jsonify({

            "error":
                "No se han podido cargar los trabajadores."

        }), 500


# ============================================================
# PLANIFICACIÓN — CAMBIOS FORZADOS
# ============================================================

@app.route("/api/cambios-forzados", methods=["DELETE"])
def eliminar_cambios_forzados():
    try:
        with engine.begin() as conn:
            resultado = conn.execute(
                text("DELETE FROM cambios_planificacion")
            )

        return jsonify({
            "ok": True,
            "mensaje": f"Se han eliminado {resultado.rowcount} cambios forzados."
        })

    except Exception as e:
        print("ERROR ELIMINANDO CAMBIOS FORZADOS:", e)
        return jsonify({
            "ok": False,
            "error": "No se han podido eliminar los cambios forzados.",
            "detalle": str(e)
        }), 500


@app.route("/api/cambios-forzados", methods=["POST"])
def crear_cambio_forzado():

    try:

        datos = request.get_json()

        fecha = datos.get("fecha")
        id_tarea = datos.get("id_tarea")
        trabajador_anterior = datos.get("trabajador_anterior")
        trabajador_nuevo = datos.get("trabajador_nuevo")
        turno = datos.get("turno")
        motivo = datos.get("motivo")

        # ----------------------------------------------------
        # Comprobar que están todos los datos
        # ----------------------------------------------------

        if not all([
            fecha,
            id_tarea,
            trabajador_anterior,
            trabajador_nuevo,
            turno is not None
        ]):

            return jsonify({
                "ok": False,
                "error": "Faltan datos para realizar el cambio."
            }), 400


        # ----------------------------------------------------
        # Obtener la versión activa
        # ----------------------------------------------------

        version = pd.read_sql("""
            SELECT id_version
            FROM planificacion_version
            WHERE activa = 1
            ORDER BY id_version DESC
            LIMIT 1
        """, engine)


        if version.empty:

            return jsonify({
                "ok": False,
                "error": "No existe una versión activa de la planificación."
            }), 400


        id_version = int(
            version.iloc[0]["id_version"]
        )


        # ----------------------------------------------------
        # Realizar el cambio
        # ----------------------------------------------------

        with engine.begin() as conn:

            # -----------------------------------------------
            # 1. Actualizar calendarización
            # -----------------------------------------------

            resultado_update = conn.execute(
                text("""
                    UPDATE calendarizacion
                    SET id_trabajador = :trabajador_nuevo
                    WHERE fecha = :fecha
                      AND id_tarea = :id_tarea
                      AND id_trabajador = :trabajador_anterior
                      AND turno = :turno
                      AND id_version = :id_version
                """),
                {
                    "fecha": fecha,
                    "id_tarea": id_tarea,
                    "trabajador_anterior": trabajador_anterior,
                    "trabajador_nuevo": trabajador_nuevo,
                    "turno": turno,
                    "id_version": id_version
                }
            )


            # -----------------------------------------------
            # Comprobar que realmente se ha actualizado
            # -----------------------------------------------

            if resultado_update.rowcount == 0:

                raise Exception(
                    "No se ha encontrado la asignación "
                    "que se quiere modificar."
                )


            # -----------------------------------------------
            # 2. Registrar el cambio forzado
            # -----------------------------------------------

            conn.execute(
                text("""
                    INSERT INTO cambios_planificacion (
                        id_version,
                        fecha,
                        fecha_asignacion,
                        id_tarea,
                        trabajador_anterior,
                        trabajador_nuevo,
                        turno,
                        motivo,
                        forzado
                    )
                    VALUES (
                        :id_version,
                        :fecha,
                        CURDATE(),
                        :id_tarea,
                        :trabajador_anterior,
                        :trabajador_nuevo,
                        :turno,
                        :motivo,
                        1
                    )
                """),
                {
                    "id_version": id_version,
                    "fecha": fecha,
                    "id_tarea": id_tarea,
                    "trabajador_anterior": trabajador_anterior,
                    "trabajador_nuevo": trabajador_nuevo,
                    "turno": turno,
                    "motivo": motivo
                }
            )


        return jsonify({
            "ok": True,
            "mensaje": "Cambio forzado realizado correctamente.",
            "id_version": id_version
        })


    except Exception as e:

        print("ERROR CAMBIO FORZADO:", e)

        return jsonify({
            "ok": False,
            "error": "No se ha podido realizar el cambio forzado.",
            "detalle": str(e)
        }), 500


    
# ============================================================
# PLANIFICACIÓN — TRABAJADORES
# ============================================================

@app.route("/api/trabajadores")
def api_trabajadores():

    trabajadores = pd.read_sql("""
        SELECT
            id_trabajador,
            CONCAT(nombre, ' ', apellidos) AS trabajador
        FROM trabajador
        WHERE activo = 1
        ORDER BY apellidos, nombre
    """, engine)

    return jsonify(
        trabajadores.to_dict(orient="records")
    )

    
# ============================================================
# CONFIGURACIÓN
# ============================================================

@app.route("/configuracion")
def configuracion():
    return render_template("configuracion.html")


@app.route("/api/configuracion", methods=["GET"])
def api_configuracion():
    try:
        return jsonify(get_all_config())
    except Exception as e:
        print("ERROR CONFIGURACIÓN:", e)
        return jsonify({
            "error": "No se han podido cargar los parámetros del modelo."
        }), 500


@app.route("/api/configuracion", methods=["POST"])
def guardar_configuracion():
    try:
        datos = request.get_json(silent=True) or {}
        config = save_all_config(datos)
        return jsonify({
            "ok": True,
            "configuracion": config
        })
    except Exception as e:
        print("ERROR GUARDAR CONFIGURACIÓN:", e)
        return jsonify({
            "error": "No se han podido guardar los parámetros del modelo."
        }), 500


if __name__ == "__main__":
    app.run(debug=True)